from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from pathlib import Path
from tempfile import NamedTemporaryFile
from threading import Lock
from typing import Any
from uuid import uuid4

from flask import Flask, jsonify, request
from openpyxl import load_workbook
from server_core.services.ranking_service import (
    compute_player_metrics as _svc_compute_player_metrics,
    is_lower_better_column as _svc_is_lower_better_column,
    normalize_player_dataset_doc as _svc_normalize_player_dataset_doc,
)


APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
STATE_PATH = DATA_DIR / "state.json"
STATE_BAK_PATH = DATA_DIR / "state.json.bak"
PLAYER_DATA_PATH = DATA_DIR / "player_dataset.json"
PLAYER_DATA_BAK_PATH = DATA_DIR / "player_dataset.json.bak"
PLAYER_DATASETS_DIR = DATA_DIR / "player_datasets"
PLAYER_DATA_INDEX_PATH = DATA_DIR / "player_datasets_index.json"
PLAYER_DATA_INDEX_BAK_PATH = DATA_DIR / "player_datasets_index.json.bak"
VERSION = 1
WRITE_LOCK = Lock()


app = Flask(__name__)


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    PLAYER_DATASETS_DIR.mkdir(parents=True, exist_ok=True)


def _corsify_response(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Methods"] = "GET,PUT,POST,OPTIONS"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return resp


@app.after_request
def _after_request(resp):
    return _corsify_response(resp)


@app.route("/api/<path:_path>", methods=["OPTIONS"])
def options_handler(_path: str):
    return _corsify_response(jsonify({"ok": True}))


def _load_doc() -> dict[str, Any] | None:
    if not STATE_PATH.exists():
        return None
    with STATE_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def _load_player_doc() -> dict[str, Any] | None:
    if not PLAYER_DATA_PATH.exists():
        return None
    with PLAYER_DATA_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def _default_player_index() -> dict[str, Any]:
    return {
        "version": VERSION,
        "updatedAt": None,
        "selectedDatasetId": "",
        "datasets": [],
    }


def _dataset_file_path(dataset_id: str) -> Path:
    return PLAYER_DATASETS_DIR / f"{dataset_id}.json"


def _load_player_index() -> dict[str, Any]:
    if PLAYER_DATA_INDEX_PATH.exists():
        with PLAYER_DATA_INDEX_PATH.open("r", encoding="utf-8") as f:
            idx = json.load(f)
            if isinstance(idx, dict):
                return {
                    "version": int(idx.get("version", VERSION)),
                    "updatedAt": idx.get("updatedAt"),
                    "selectedDatasetId": str(idx.get("selectedDatasetId") or ""),
                    "datasets": idx.get("datasets") if isinstance(idx.get("datasets"), list) else [],
                }
    legacy = _load_player_doc()
    if legacy and isinstance(legacy, dict) and legacy.get("players"):
        dataset_id = "legacy"
        _write_dataset_doc(dataset_id, legacy)
        imported_at = legacy.get("updatedAt") or _iso_now()
        idx = {
            "version": VERSION,
            "updatedAt": imported_at,
            "selectedDatasetId": dataset_id,
            "datasets": [
                {
                    "id": dataset_id,
                    "name": f"legacy-{imported_at[:19]}",
                    "updatedAt": imported_at,
                    "playerCount": len(legacy.get("players", [])),
                    "numericColumnCount": len((legacy.get("schema") or {}).get("numericColumns", [])),
                    "sourceFile": (legacy.get("source") or {}).get("filename", "legacy"),
                }
            ],
        }
        _atomic_write_player_index(idx)
        return idx
    return _default_player_index()


def _load_dataset_doc(dataset_id: str) -> dict[str, Any] | None:
    path = _dataset_file_path(dataset_id)
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _atomic_write_json(path: Path, bak_path: Path, prefix: str, doc: dict[str, Any]) -> None:
    _ensure_data_dir()
    data = json.dumps(doc, ensure_ascii=False, indent=2)
    with WRITE_LOCK:
        if path.exists():
            bak_path.write_text(path.read_text(encoding="utf-8"), encoding="utf-8")
        with NamedTemporaryFile("w", encoding="utf-8", delete=False, dir=path.parent, prefix=prefix, suffix=".tmp") as tf:
            tf.write(data)
            tmp_path = Path(tf.name)
        os.replace(tmp_path, path)


def _validate_payload(payload: Any) -> tuple[bool, str]:
    if not isinstance(payload, dict):
        return False, "payload must be object"
    for key in ("draft", "presets", "selectedPresetId"):
        if key not in payload:
            return False, f"missing key: {key}"
    if not isinstance(payload["presets"], list):
        return False, "presets must be array"
    if not isinstance(payload["selectedPresetId"], str):
        return False, "selectedPresetId must be string"
    return True, ""


def _atomic_write_doc(doc: dict[str, Any]) -> None:
    _ensure_data_dir()
    data = json.dumps(doc, ensure_ascii=False, indent=2)
    with WRITE_LOCK:
        if STATE_PATH.exists():
            STATE_BAK_PATH.write_text(STATE_PATH.read_text(encoding="utf-8"), encoding="utf-8")
        with NamedTemporaryFile("w", encoding="utf-8", delete=False, dir=DATA_DIR, prefix="state_", suffix=".tmp") as tf:
            tf.write(data)
            tmp_path = Path(tf.name)
        os.replace(tmp_path, STATE_PATH)


def _atomic_write_player_doc(doc: dict[str, Any]) -> None:
    _atomic_write_json(PLAYER_DATA_PATH, PLAYER_DATA_BAK_PATH, "player_data_", doc)


def _write_dataset_doc(dataset_id: str, doc: dict[str, Any]) -> None:
    path = _dataset_file_path(dataset_id)
    bak_path = PLAYER_DATASETS_DIR / f"{dataset_id}.bak.json"
    _atomic_write_json(path, bak_path, f"dataset_{dataset_id}_", doc)


def _atomic_write_player_index(doc: dict[str, Any]) -> None:
    _atomic_write_json(PLAYER_DATA_INDEX_PATH, PLAYER_DATA_INDEX_BAK_PATH, "player_datasets_index_", doc)


def _build_doc(payload: dict[str, Any]) -> dict[str, Any]:
    return {
        "version": VERSION,
        "updatedAt": _iso_now(),
        "data": payload,
    }


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = text.replace(",", "")
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _to_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def _normalize_header_name(value: Any) -> str:
    text = str(value or "").strip().lower().replace("_", " ")
    return " ".join(text.split())


def _pick_name_column(headers: list[str]) -> tuple[int, str]:
    candidates = {"player", "name", "player name", "姓名", "球员", "球员姓名"}
    for idx, header in enumerate(headers):
        normalized = _normalize_header_name(header)
        if normalized in candidates:
            return idx, header
    return -1, ""


def _pick_team_column(headers: list[str]) -> tuple[int, str]:
    exact_candidates = {"team", "club", "squad", "球队", "俱乐部"}
    for idx, header in enumerate(headers):
        normalized = _normalize_header_name(header)
        if normalized in exact_candidates:
            return idx, header

    keyword_candidates = ("team", "club", "squad", "球队", "俱乐部")
    for idx, header in enumerate(headers):
        normalized = _normalize_header_name(header)
        if any(keyword in normalized for keyword in keyword_candidates):
            return idx, header
    return -1, ""


def _is_lower_better_column(column_name: str) -> bool:
    return _svc_is_lower_better_column(column_name)


def _compute_player_metrics(
    players: list[dict[str, Any]],
    candidate_numeric_cols: list[str],
) -> tuple[list[str], list[str]]:
    return _svc_compute_player_metrics(
        players,
        candidate_numeric_cols,
        to_float_fn=_to_float,
        is_lower_better_fn=_is_lower_better_column,
    )


def _normalize_player_dataset_doc(doc: dict[str, Any]) -> dict[str, Any]:
    return _svc_normalize_player_dataset_doc(
        doc,
        to_float_fn=_to_float,
        is_lower_better_fn=_is_lower_better_column,
    )


def _make_player_id(player_name: str, used_ids: dict[str, int]) -> str:
    base = "".join(ch.lower() if ch.isalnum() else "_" for ch in player_name).strip("_")
    if not base:
        base = "player"
    if base not in used_ids:
        used_ids[base] = 1
        return base
    used_ids[base] += 1
    return f"{base}_{used_ids[base]}"


def _build_player_columns(player: dict[str, Any], schema: dict[str, Any]) -> list[dict[str, Any]]:
    numeric_cols = set(schema.get("numericColumns", []))
    all_cols = schema.get("allColumns", [])
    raw = player.get("raw", {})
    metrics = player.get("metrics", {})
    rows = []
    for col in all_cols:
        metric = metrics.get(col, {})
        rows.append(
            {
                "column": col,
                "value": raw.get(col, ""),
                "rank": metric.get("rank"),
                "percentile": metric.get("percentile"),
                "isNumeric": col in numeric_cols,
            }
        )
    return rows


def _resolve_dataset(index: dict[str, Any], requested_dataset_id: str) -> tuple[str, dict[str, Any] | None]:
    dataset_id = requested_dataset_id or str(index.get("selectedDatasetId") or "")
    if not dataset_id:
        return "", None
    return dataset_id, _load_dataset_doc(dataset_id)


@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"ok": True, "ts": _iso_now()})


@app.route("/api/state", methods=["GET"])
def get_state():
    try:
        doc = _load_doc()
        if doc is None:
            return jsonify({"ok": True, "version": VERSION, "updatedAt": None, "data": None})
        return jsonify(
            {
                "ok": True,
                "version": int(doc.get("version", VERSION)),
                "updatedAt": doc.get("updatedAt"),
                "data": doc.get("data"),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


@app.route("/api/state", methods=["PUT"])
def put_state():
    payload = request.get_json(silent=True)
    ok, message = _validate_payload(payload)
    if not ok:
        return jsonify({"ok": False, "error": message}), 400

    try:
        doc = _build_doc(payload)
        _atomic_write_doc(doc)
        return jsonify({"ok": True, "updatedAt": doc["updatedAt"]})
    except Exception as exc:
        return jsonify({"ok": False, "error": f"write failed: {exc}"}), 500


@app.route("/api/migrate-from-local", methods=["POST"])
def migrate_from_local():
    payload = request.get_json(silent=True)
    ok, message = _validate_payload(payload)
    if not ok:
        return jsonify({"ok": False, "error": message}), 400

    try:
        existing = _load_doc()
        if existing and existing.get("data"):
            return jsonify({"ok": True, "migrated": False, "skipped": True})
        doc = _build_doc(payload)
        doc["migrationSource"] = "localStorage"
        _atomic_write_doc(doc)
        return jsonify({"ok": True, "migrated": True, "skipped": False, "updatedAt": doc["updatedAt"]})
    except Exception as exc:
        return jsonify({"ok": False, "error": f"migrate failed: {exc}"}), 500


@app.route("/api/player-data/import-excel", methods=["POST"])
def import_player_data_excel():
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "only .xlsx is supported"}), 400

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"invalid excel file: {exc}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return jsonify({"ok": False, "error": "excel must contain header and data rows"}), 400

    headers_raw = [_to_cell_value(x) for x in rows[0]]
    headers = [str(x).strip() for x in headers_raw]
    headers_lower = [h.lower() for h in headers]
    if "player" not in headers_lower:
        return jsonify({"ok": False, "error": "missing required column: player"}), 400
    player_idx = headers_lower.index("player")

    parsed_players: list[dict[str, Any]] = []
    used_ids: dict[str, int] = {}
    candidate_numeric_cols = [h for h in headers if h and h.lower() != "player"]

    for row_idx, excel_row in enumerate(rows[1:], start=1):
        cells = list(excel_row) + [None] * max(0, len(headers) - len(excel_row))
        raw = {}
        for i, col in enumerate(headers):
            raw[col] = _to_cell_value(cells[i] if i < len(cells) else None)

        player_name = str(raw.get(headers[player_idx], "")).strip()
        if not player_name:
            continue

        player = {
            "id": _make_player_id(player_name, used_ids),
            "player": player_name,
            "raw": raw,
            "metrics": {},
            "_numeric": {},
            "_rowIndex": row_idx,
        }

        parsed_players.append(player)

    if not parsed_players:
        return jsonify({"ok": False, "error": "no valid player rows found"}), 400

    numeric_columns, lower_better_columns = _compute_player_metrics(parsed_players, candidate_numeric_cols)
    for player in parsed_players:
        player.pop("_numeric", None)
        player.pop("_rowIndex", None)

    doc = {
        "version": VERSION,
        "updatedAt": _iso_now(),
        "source": {
            "filename": file.filename,
            "sheet": ws.title,
            "rowCount": len(parsed_players),
        },
        "schema": {
            "playerColumn": "player",
            "numericColumns": numeric_columns,
            "lowerBetterColumns": lower_better_columns,
            "allColumns": headers,
        },
        "players": parsed_players,
    }
    dataset_id = f"ds_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"
    doc["datasetId"] = dataset_id

    try:
        _write_dataset_doc(dataset_id, doc)
        index = _load_player_index()
        datasets = [d for d in index.get("datasets", []) if isinstance(d, dict)]
        datasets.insert(
            0,
            {
                "id": dataset_id,
                "name": f"{file.filename} ({doc['updatedAt'][:19]})",
                "updatedAt": doc["updatedAt"],
                "playerCount": len(parsed_players),
                "numericColumnCount": len(numeric_columns),
                "sourceFile": file.filename,
            },
        )
        index["datasets"] = datasets
        index["selectedDatasetId"] = dataset_id
        index["updatedAt"] = doc["updatedAt"]
        _atomic_write_player_index(index)
        _atomic_write_player_doc(doc)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"write failed: {exc}"}), 500

    return jsonify(
        {
            "ok": True,
            "datasetId": dataset_id,
            "updatedAt": doc["updatedAt"],
            "playerCount": len(parsed_players),
            "numericColumnCount": len(numeric_columns),
        }
    )
@app.route("/api/name-mapping/import-excel", methods=["POST"])
def import_name_mapping_excel():
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "only .xlsx is supported"}), 400

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"invalid excel file: {exc}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return jsonify({"ok": False, "error": "excel must contain header and data rows"}), 400

    headers = [str(_to_cell_value(x)).strip() for x in rows[0]]
    name_col_idx, name_col = _pick_name_column(headers)
    if name_col_idx < 0:
        return jsonify({"ok": False, "error": "missing required name column (Player/Name)"}), 400
    team_col_idx, team_col = _pick_team_column(headers)

    names: list[str] = []
    items: list[dict[str, str]] = []
    seen: set[str] = set()
    for excel_row in rows[1:]:
        cells = list(excel_row)
        name_value = cells[name_col_idx] if name_col_idx < len(cells) else None
        name = str(_to_cell_value(name_value)).strip()
        if not name:
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        names.append(name)
        team_value = cells[team_col_idx] if team_col_idx >= 0 and team_col_idx < len(cells) else None
        team_en = str(_to_cell_value(team_value)).strip()
        items.append({"name": name, "teamEn": team_en})

    if not names:
        return jsonify({"ok": False, "error": "no valid name rows found"}), 400

    return jsonify(
        {
            "ok": True,
            "names": names,
            "items": items,
            "count": len(names),
            "sheet": ws.title,
            "column": name_col,
            "teamColumn": team_col,
        }
    )
@app.route("/api/project-mapping/import-excel", methods=["POST"])
def import_project_mapping_excel():
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "only .xlsx is supported"}), 400

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"invalid excel file: {exc}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True, max_row=1))
    wb.close()
    if not rows:
        return jsonify({"ok": False, "error": "excel must contain at least one header row"}), 400

    seen: set[str] = set()
    columns: list[str] = []
    for cell in list(rows[0]):
        header = str(_to_cell_value(cell)).strip()
        if not header:
            continue
        key = _normalize_header_name(header)
        if not key or key in seen:
            continue
        seen.add(key)
        columns.append(header)

    if not columns:
        return jsonify({"ok": False, "error": "no valid header columns found"}), 400

    return jsonify(
        {
            "ok": True,
            "sheet": ws.title,
            "columns": columns,
            "count": len(columns),
        }
    )
@app.route("/api/player-data", methods=["GET"])
def get_player_data():
    try:
        index = _load_player_index()
        dataset_id, doc = _resolve_dataset(index, str(request.args.get("datasetId") or ""))
        if doc is None:
            return jsonify({"ok": True, "data": None, "updatedAt": None, "datasetId": dataset_id, "selectedDatasetId": index.get("selectedDatasetId")})
        doc = _normalize_player_dataset_doc(doc)
        return jsonify(
            {
                "ok": True,
                "updatedAt": doc.get("updatedAt"),
                "datasetId": dataset_id,
                "selectedDatasetId": index.get("selectedDatasetId"),
                "data": doc,
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


@app.route("/api/player-data/datasets", methods=["GET"])
def get_player_datasets():
    try:
        index = _load_player_index()
        return jsonify(
            {
                "ok": True,
                "datasets": index.get("datasets", []),
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "updatedAt": index.get("updatedAt"),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


@app.route("/api/player-data/datasets/<dataset_id>", methods=["DELETE"])
def delete_player_dataset(dataset_id: str):
    try:
        index = _load_player_index()
        datasets = [d for d in index.get("datasets", []) if isinstance(d, dict)]
        matched = next((d for d in datasets if d.get("id") == dataset_id), None)
        if matched is None:
            return jsonify({"ok": False, "error": "dataset not found"}), 404

        path = _dataset_file_path(dataset_id)
        if path.exists():
            path.unlink()

        remaining = [d for d in datasets if d.get("id") != dataset_id]
        index["datasets"] = remaining
        if index.get("selectedDatasetId") == dataset_id:
            index["selectedDatasetId"] = remaining[0].get("id") if remaining else ""
        index["updatedAt"] = _iso_now()
        _atomic_write_player_index(index)
        return jsonify({"ok": True, "deletedDatasetId": dataset_id, "selectedDatasetId": index.get("selectedDatasetId", ""), "datasets": remaining})
    except Exception as exc:
        return jsonify({"ok": False, "error": f"delete failed: {exc}"}), 500


@app.route("/api/player-data/players", methods=["GET"])
def get_player_list():
    try:
        index = _load_player_index()
        dataset_id, doc = _resolve_dataset(index, str(request.args.get("datasetId") or ""))
        if doc is None:
            return jsonify({"ok": True, "players": [], "playerCount": 0, "updatedAt": None, "numericColumns": [], "datasetId": dataset_id, "selectedDatasetId": index.get("selectedDatasetId", "")})
        doc = _normalize_player_dataset_doc(doc)
        players = doc.get("players", [])
        items = [{"id": p.get("id"), "player": p.get("player")} for p in players if p.get("id") and p.get("player")]
        schema = doc.get("schema", {})
        return jsonify(
            {
                "ok": True,
                "players": items,
                "playerCount": len(items),
                "datasetId": dataset_id,
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "updatedAt": doc.get("updatedAt"),
                "numericColumns": schema.get("numericColumns", []),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


@app.route("/api/player-data/player/<player_id>", methods=["GET"])
def get_player_detail(player_id: str):
    try:
        index = _load_player_index()
        dataset_id, doc = _resolve_dataset(index, str(request.args.get("datasetId") or ""))
        if doc is None:
            return jsonify({"ok": False, "error": "player dataset not found"}), 404
        doc = _normalize_player_dataset_doc(doc)
        players = doc.get("players", [])
        selected = next((p for p in players if p.get("id") == player_id), None)
        if selected is None:
            return jsonify({"ok": False, "error": "player not found"}), 404
        schema = doc.get("schema", {})
        return jsonify(
            {
                "ok": True,
                "player": {
                    "id": selected.get("id"),
                    "player": selected.get("player"),
                    "columns": _build_player_columns(selected, schema),
                },
                "datasetId": dataset_id,
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "updatedAt": doc.get("updatedAt"),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


if __name__ == "__main__":
    _ensure_data_dir()
    app.run(host="127.0.0.1", port=8787, debug=False)
