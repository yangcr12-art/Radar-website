from __future__ import annotations

from io import BytesIO

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook, load_workbook

from server_core.services.player_dataset_store import pick_name_column, pick_team_column, to_cell_value, normalize_header_name


mapping_import_bp = Blueprint("mapping_import_api", __name__)


def _table_rows_from_sheet(rows, required_headers: list[str]) -> tuple[list[dict[str, str]] | None, str]:
    if len(rows) < 2:
        return None, ""
    headers = [str(to_cell_value(x)).strip() for x in rows[0]]
    if not all(header in headers for header in required_headers):
        return None, ""
    index_by_header = {header: headers.index(header) for header in required_headers}
    items: list[dict[str, str]] = []
    seen: set[str] = set()
    for excel_row in rows[1:]:
        cells = list(excel_row)
        item = {
            header: str(to_cell_value(cells[index_by_header[header]])).strip() if index_by_header[header] < len(cells) else ""
            for header in required_headers
        }
        en = str(item.get("English") or "").strip()
        key = en.lower()
        if not any(item.values()):
            continue
        if not en or key in seen:
            continue
        seen.add(key)
        items.append(item)
    return items, headers[0] if headers else ""


def _build_excel_response(filename: str, headers: list[str], rows: list[list[str]]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@mapping_import_bp.route("/api/name-mapping/import-excel", methods=["POST"])
def import_name_mapping_excel():
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "only .xlsx is supported"}), 400
    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"invalid excel file: {exc}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    table_items, _ = _table_rows_from_sheet(rows, ["English", "中文翻译", "球队"])
    if table_items is not None:
        return jsonify(
            {
                "ok": True,
                "items": [{"en": item["English"], "zh": item["中文翻译"], "team": item["球队"]} for item in table_items],
                "count": len(table_items),
                "sheet": ws.title,
            }
        )

    if len(rows) < 2:
        return jsonify({"ok": False, "error": "excel must contain header and data rows"}), 400

    headers = [str(to_cell_value(x)).strip() for x in rows[0]]
    name_col_idx, name_col = pick_name_column(headers)
    if name_col_idx < 0:
        return jsonify({"ok": False, "error": "missing required name column (Player/Name)"}), 400
    team_col_idx, team_col = pick_team_column(headers)

    names: list[str] = []
    items: list[dict[str, str]] = []
    seen: set[str] = set()
    for excel_row in rows[1:]:
        cells = list(excel_row)
        name_value = cells[name_col_idx] if name_col_idx < len(cells) else None
        name = str(to_cell_value(name_value)).strip()
        if not name:
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        names.append(name)
        team_value = cells[team_col_idx] if team_col_idx >= 0 and team_col_idx < len(cells) else None
        team_en = str(to_cell_value(team_value)).strip()
        items.append({"en": name, "zh": "", "team": team_en})

    if not names:
        return jsonify({"ok": False, "error": "no valid name rows found"}), 400

    return jsonify(
        {
            "ok": True,
            "names": names,
            "items": items,
            "count": len(names),
            "sheet": ws.title,
            "column": name_col,
            "teamColumn": team_col,
        }
    )


@mapping_import_bp.route("/api/name-mapping/export-excel", methods=["POST"])
def export_name_mapping_excel():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows")
    if not isinstance(rows, list):
        return jsonify({"ok": False, "error": "rows must be array"}), 400
    data_rows = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        data_rows.append([
            str(item.get("en") or "").strip(),
            str(item.get("zh") or "").strip(),
            str(item.get("team") or "").strip(),
        ])
    return _build_excel_response("name_mapping.xlsx", ["English", "中文翻译", "球队"], data_rows)


@mapping_import_bp.route("/api/project-mapping/import-excel", methods=["POST"])
def import_project_mapping_excel():
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "only .xlsx is supported"}), 400
    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"invalid excel file: {exc}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return jsonify({"ok": False, "error": "excel must contain at least one header row"}), 400

    table_items, _ = _table_rows_from_sheet(rows, ["English", "中文翻译", "group"])
    if table_items is not None:
        return jsonify(
            {
                "ok": True,
                "items": [{"en": item["English"], "zh": item["中文翻译"], "group": item["group"]} for item in table_items],
                "count": len(table_items),
                "sheet": ws.title,
            }
        )

    seen: set[str] = set()
    columns: list[str] = []
    for cell in list(rows[0]):
        header = str(to_cell_value(cell)).strip()
        if not header:
            continue
        key = normalize_header_name(header)
        if not key or key in seen:
            continue
        seen.add(key)
        columns.append(header)

    if not columns:
        return jsonify({"ok": False, "error": "no valid header columns found"}), 400

    return jsonify(
        {
            "ok": True,
            "sheet": ws.title,
            "columns": columns,
            "count": len(columns),
        }
    )


@mapping_import_bp.route("/api/project-mapping/export-excel", methods=["POST"])
def export_project_mapping_excel():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows")
    if not isinstance(rows, list):
        return jsonify({"ok": False, "error": "rows must be array"}), 400
    data_rows = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        data_rows.append([
            str(item.get("en") or "").strip(),
            str(item.get("zh") or "").strip(),
            str(item.get("group") or "").strip(),
        ])
    return _build_excel_response("project_mapping.xlsx", ["English", "中文翻译", "group"], data_rows)


@mapping_import_bp.route("/api/team-mapping/import-excel", methods=["POST"])
def import_team_mapping_excel():
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "only .xlsx is supported"}), 400
    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"invalid excel file: {exc}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    table_items, _ = _table_rows_from_sheet(rows, ["English", "中文翻译", "color", "shape", "logoFileName"])
    if table_items is None:
        return jsonify({"ok": False, "error": "excel must contain headers: English, 中文翻译, color, shape, logoFileName"}), 400

    return jsonify(
        {
            "ok": True,
            "items": [
                {
                    "en": item["English"],
                    "zh": item["中文翻译"],
                    "color": item["color"],
                    "shape": item["shape"],
                    "logoFileName": item["logoFileName"],
                }
                for item in table_items
            ],
            "count": len(table_items),
            "sheet": ws.title,
        }
    )


@mapping_import_bp.route("/api/team-mapping/export-excel", methods=["POST"])
def export_team_mapping_excel():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows")
    if not isinstance(rows, list):
        return jsonify({"ok": False, "error": "rows must be array"}), 400
    data_rows = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        data_rows.append([
            str(item.get("en") or "").strip(),
            str(item.get("zh") or "").strip(),
            str(item.get("color") or "").strip(),
            str(item.get("shape") or "").strip(),
            str(item.get("logoFileName") or "").strip(),
        ])
    return _build_excel_response("team_mapping.xlsx", ["English", "中文翻译", "color", "shape", "logoFileName"], data_rows)
