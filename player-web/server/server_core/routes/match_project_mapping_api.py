from __future__ import annotations

from typing import Any

from flask import Blueprint, jsonify, request
from openpyxl import load_workbook


match_project_mapping_bp = Blueprint("match_project_mapping_api", __name__)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_tokens(text: str) -> list[str]:
    base = str(text or "").replace("／", "/").strip()
    if not base:
        return []
    tokens = [item.strip() for item in base.split("/") if item.strip()]
    return tokens


def _infer_group(en_name: str, zh_name: str) -> str:
    text = f"{en_name} {zh_name}".lower()
    if any(k in text for k in ["pass", "传球", "长传", "直塞", "妙传"]):
        return "传球"
    if any(k in text for k in ["recover", "loss", "duel", "defen", "interception", "foul", "夺回", "丢失", "对抗", "防守", "拦截", "犯规"]):
        return "防守"
    if any(k in text for k in ["shot", "goal", "xg", "corner", "free kick", "penalt", "射门", "进球", "角球", "任意球", "点球"]):
        return "进攻"
    return "其他"


def _split_special_cases(en_anchor: str, zh_anchor: str, span: int) -> list[tuple[str, str]]:
    en_norm = en_anchor.lower().replace(" ", "")
    zh_norm = zh_anchor.replace(" ", "")

    if span == 3 and ("shots/ontarget" in en_norm or "射门/射正" in zh_norm):
        return [
            ("Shots", "射门数"),
            ("Shots on target", "射正数"),
            ("Shots on target, %", "射正率"),
        ]

    if span == 4 and ("recoveries/low/medium/high" in en_norm or "夺回球权/低位/中位/高位" in zh_norm):
        return [
            ("Recoveries", "夺回球权总数"),
            ("Recoveries - Low", "低位夺回球权数"),
            ("Recoveries - Medium", "中位夺回球权数"),
            ("Recoveries - High", "高位夺回球权数"),
        ]

    if span == 4 and ("losses/low/medium/high" in en_norm or "丢失球权/低位/中位/高位" in zh_norm):
        return [
            ("Losses", "丢失球权总数"),
            ("Losses - Low", "低位丢失球权数"),
            ("Losses - Medium", "中位丢失球权数"),
            ("Losses - High", "高位丢失球权数"),
        ]

    if span == 3 and ("penaltyareaentries" in en_norm or "攻入禁区" in zh_norm):
        return [
            ("Penalty area entries", "攻入禁区"),
            ("Penalty area entries - Runs", "带球攻入禁区"),
            ("Penalty area entries - Crosses", "传中入禁区"),
        ]

    return []


def _build_span_names(en_anchor: str, zh_anchor: str, span: int) -> list[tuple[str, str]]:
    if span <= 1:
        en = _cell_text(en_anchor)
        zh = _cell_text(zh_anchor)
        return [(en, zh or en)]

    special = _split_special_cases(_cell_text(en_anchor), _cell_text(zh_anchor), span)
    if special:
        return special

    en_tokens = _split_tokens(_cell_text(en_anchor))
    zh_tokens = _split_tokens(_cell_text(zh_anchor))
    if span == len(en_tokens) and span == len(zh_tokens) and span > 0:
        return list(zip(en_tokens, zh_tokens))

    if span == 4 and len(en_tokens) == 4 and len(zh_tokens) == 4:
        return list(zip(en_tokens, zh_tokens))

    if span == 3 and len(en_tokens) == 2 and len(zh_tokens) == 2:
        base_en, qual_en = en_tokens
        base_zh, qual_zh = zh_tokens
        return [
            (base_en, f"{base_zh}数"),
            (f"{base_en}, {qual_en}", f"{qual_zh}数"),
            (f"{base_en}, {qual_en} %", f"{qual_zh}率"),
        ]

    if span == len(en_tokens) + 1 and len(en_tokens) >= 1:
        rows: list[tuple[str, str]] = []
        for idx in range(len(en_tokens)):
            en = en_tokens[idx]
            zh = zh_tokens[idx] if idx < len(zh_tokens) else en
            rows.append((en, zh))
        tail_en = en_tokens[-1]
        tail_zh = zh_tokens[-1] if zh_tokens else tail_en
        rows.append((f"{tail_en}, %", f"{tail_zh}率"))
        return rows

    en = _cell_text(en_anchor) or "Metric"
    zh = _cell_text(zh_anchor) or en
    return [(f"{en} #{idx + 1}", f"{zh}{idx + 1}") for idx in range(span)]


def _normalize_loss_levels(en: str, zh: str, prev_anchor_en: str) -> tuple[str, str]:
    en_text = _cell_text(en)
    zh_text = _cell_text(zh)
    prev = _cell_text(prev_anchor_en).lower()
    if prev != "losses":
        return en_text, zh_text
    level_map = {
        "low": ("Losses - Low", "低位丢失球权数"),
        "medium": ("Losses - Medium", "中位丢失球权数"),
        "high": ("Losses - High", "高位丢失球权数"),
    }
    mapped = level_map.get(en_text.lower())
    if mapped:
        return mapped
    return en_text, zh_text


def _normalize_penalty_entries(en: str, zh: str) -> tuple[str, str]:
    en_text = _cell_text(en)
    en_norm = en_text.lower().replace(" ", "")
    if "penaltyareaentries(runs,crosses)%" in en_norm:
        return "Penalty area entries - Crosses", "传中入禁区"
    if "penaltyareaentries(runs,crosses)" in en_norm:
        return "Penalty area entries - Runs", "带球攻入禁区"
    if "penaltyareaentries(runs" in en_norm:
        return "Penalty area entries", "攻入禁区"
    return en_text, _cell_text(zh)


@match_project_mapping_bp.route("/api/match-project-mapping/import-excel", methods=["POST"])
def import_match_project_mapping_excel():
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "only .xlsx is supported"}), 400

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"invalid excel file: {exc}"}), 400

    ws = wb.active
    maxc = ws.max_column or 0
    if maxc <= 0:
        wb.close()
        return jsonify({"ok": False, "error": "empty sheet"}), 400

    en_row = [_cell_text(ws.cell(1, c).value) for c in range(1, maxc + 1)]
    zh_row = [_cell_text(ws.cell(2, c).value) for c in range(1, maxc + 1)]
    wb.close()

    anchors = [idx for idx, value in enumerate(en_row) if value]
    if not anchors:
        return jsonify({"ok": False, "error": "row 1 has no english anchors"}), 400

    items: list[dict[str, str]] = []
    seen: set[str] = set()
    warnings: list[str] = []
    prev_anchor_en = ""
    for i, anchor_idx in enumerate(anchors):
        next_anchor = anchors[i + 1] if i + 1 < len(anchors) else maxc
        span = max(1, next_anchor - anchor_idx)
        en_anchor = en_row[anchor_idx]
        zh_anchor = zh_row[anchor_idx] or en_anchor
        pairs = _build_span_names(en_anchor, zh_anchor, span)
        if len(pairs) != span:
            warnings.append(f"anchor '{en_anchor}' span={span} fallback-used")
            pairs = (pairs + [(f"{en_anchor} #{n + 1}", f"{zh_anchor}{n + 1}") for n in range(span)])[:span]
        for en_name, zh_name in pairs:
            en = _cell_text(en_name)
            zh = _cell_text(zh_name) or en
            en, zh = _normalize_loss_levels(en, zh, prev_anchor_en)
            en, zh = _normalize_penalty_entries(en, zh)
            if not en:
                continue
            key = en.lower()
            if key in seen:
                continue
            seen.add(key)
            items.append({"en": en, "zh": zh, "group": _infer_group(en, zh)})
        prev_anchor_en = _cell_text(en_anchor)

    return jsonify(
        {
            "ok": True,
            "sheet": ws.title if ws else "",
            "count": len(items),
            "items": items,
            "warnings": warnings,
        }
    )
