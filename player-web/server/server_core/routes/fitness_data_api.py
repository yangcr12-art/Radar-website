from __future__ import annotations

import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from tempfile import NamedTemporaryFile
from threading import Lock
from typing import Any
from uuid import uuid4

from flask import Blueprint, jsonify, request
from openpyxl import load_workbook
from server_core.services.auth_config import get_primary_login_username
from server_core.services.session_auth import get_authenticated_username
from server_core.services.user_storage import ensure_user_data_dir, user_data_file, user_data_subdir

VERSION = 1
WRITE_LOCK = Lock()
fitness_data_bp = Blueprint("fitness_data_api", __name__)


def _resolve_username() -> str:
    return get_authenticated_username(get_primary_login_username())


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ensure_data_dir() -> None:
    ensure_user_data_dir(_resolve_username())
    user_data_subdir(_resolve_username(), "fitness_datasets")


def _fitness_datasets_dir() -> Path:
    return user_data_subdir(_resolve_username(), "fitness_datasets")


def _fitness_data_index_path() -> Path:
    return user_data_file(_resolve_username(), "fitness_datasets_index.json")


def _fitness_data_index_bak_path() -> Path:
    return user_data_file(_resolve_username(), "fitness_datasets_index.json.bak")


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "").replace("%", "")
    try:
        return float(text)
    except ValueError:
        return None


def _to_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def _normalize_header_name(value: Any) -> str:
    text = str(value or "").strip().lower().replace("_", " ")
    return " ".join(text.split())


def _normalize_token(value: Any) -> str:
    return str(value or "").strip().lower()


def _normalize_player_overview_side(value: Any) -> str:
    token = _normalize_token(value)
    return "away" if token == "away" else "home"


def _pick_team_column(headers: list[str]) -> tuple[int, str]:
    exact_candidates = {"team", "club", "squad", "球队", "俱乐部"}
    for idx, header in enumerate(headers):
        normalized = _normalize_header_name(header)
        if normalized in exact_candidates:
            return idx, header

    keyword_candidates = ("team", "club", "squad", "球队", "俱乐部")
    for idx, header in enumerate(headers):
        normalized = _normalize_header_name(header)
        if any(keyword in normalized for keyword in keyword_candidates):
            return idx, header

    return (-1, "")


def _pick_player_column(headers: list[str]) -> tuple[int, str]:
    exact_candidates = {"player", "name", "player name", "姓名", "球员", "球员姓名"}
    for idx, header in enumerate(headers):
        normalized = _normalize_header_name(header)
        if normalized in exact_candidates:
            return idx, header
    return (-1, "")


def _looks_like_numbered_player_name(text: str) -> bool:
    value = str(text or "").strip()
    if not value:
        return False
    return re.match(r"^\d+\s*-\s*.+$", value) is not None


def _normalize_headers(headers: list[str]) -> list[str]:
    normalized: list[str] = []
    used: dict[str, int] = {}
    for idx, header in enumerate(headers):
        base = str(header or "").strip() or f"column_{idx + 1}"
        key = base
        if key in used:
            used[key] += 1
            key = f"{key}_{used[key]}"
        else:
            used[key] = 1
        normalized.append(key)
    return normalized


def _pick_player_column_by_values(rows: list[list[Any]], headers: list[str]) -> tuple[int, str]:
    if len(rows) < 2 or len(headers) == 0:
        return (-1, "")

    sample_rows = rows[1 : min(len(rows), 121)]
    if len(sample_rows) == 0:
        return (-1, "")

    candidate_indices: list[int] = []
    if len(headers) > 1:
        candidate_indices.append(1)
    candidate_indices.extend([idx for idx in range(min(len(headers), 6)) if idx not in candidate_indices])

    best_idx = -1
    best_score = float("-inf")

    for idx in candidate_indices:
        non_empty = 0
        text_count = 0
        numeric_count = 0
        numbered_name_count = 0
        for row in sample_rows:
            cell = _to_cell_value(row[idx] if idx < len(row) else None)
            text = str(cell).strip()
            if not text:
                continue
            non_empty += 1
            if _to_float(text) is not None:
                numeric_count += 1
            else:
                text_count += 1
            if _looks_like_numbered_player_name(text):
                numbered_name_count += 1

        if non_empty < 2:
            continue

        text_ratio = text_count / non_empty
        numeric_ratio = numeric_count / non_empty
        if text_ratio < 0.5:
            continue

        score = (text_count * 1.0) + (numbered_name_count * 2.5) - (numeric_count * 1.2)
        if idx == 1:
            score += 1.5
        if numeric_ratio > 0.4:
            score -= 3.0

        if score > best_score:
            best_score = score
            best_idx = idx

    if best_idx < 0:
        return (-1, "")
    return best_idx, headers[best_idx]


def _score_player_doc(doc: dict[str, Any] | None) -> int:
    if not doc or not isinstance(doc, dict):
        return -1
    players = doc.get("players") if isinstance(doc.get("players"), list) else []
    metrics = doc.get("numericColumns") if isinstance(doc.get("numericColumns"), list) else []
    player_count = len(players)
    metric_count = len(metrics)
    if player_count <= 0 or metric_count <= 0:
        return -1
    return player_count * metric_count


def _make_id(name: str, used_ids: dict[str, int], fallback: str) -> str:
    base = "".join(ch.lower() if ch.isalnum() else "_" for ch in name).strip("_")
    if not base:
        base = fallback
    if base not in used_ids:
        used_ids[base] = 1
        return base
    used_ids[base] += 1
    return f"{base}_{used_ids[base]}"


def _default_index() -> dict[str, Any]:
    return {
        "version": VERSION,
        "updatedAt": None,
        "selectedDatasetId": "",
        "datasets": [],
    }


def _dataset_file_path(dataset_id: str) -> Path:
    return _fitness_datasets_dir() / f"{dataset_id}.json"


def _atomic_write_json(path: Path, bak_path: Path, prefix: str, doc: dict[str, Any]) -> None:
    _ensure_data_dir()
    data = json.dumps(doc, ensure_ascii=False, indent=2)
    with WRITE_LOCK:
        if path.exists():
            bak_path.write_text(path.read_text(encoding="utf-8"), encoding="utf-8")
        with NamedTemporaryFile("w", encoding="utf-8", delete=False, dir=path.parent, prefix=prefix, suffix=".tmp") as tf:
            tf.write(data)
            tmp_path = Path(tf.name)
        os.replace(tmp_path, path)


def _load_index() -> dict[str, Any]:
    fitness_data_index_path = _fitness_data_index_path()
    if fitness_data_index_path.exists():
        with fitness_data_index_path.open("r", encoding="utf-8") as f:
            idx = json.load(f)
            if isinstance(idx, dict):
                return {
                    "version": int(idx.get("version", VERSION)),
                    "updatedAt": idx.get("updatedAt"),
                    "selectedDatasetId": str(idx.get("selectedDatasetId") or ""),
                    "datasets": idx.get("datasets") if isinstance(idx.get("datasets"), list) else [],
                }
    return _default_index()


def _load_dataset_doc(dataset_id: str) -> dict[str, Any] | None:
    path = _dataset_file_path(dataset_id)
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _write_dataset_doc(dataset_id: str, doc: dict[str, Any]) -> None:
    path = _dataset_file_path(dataset_id)
    bak_path = _fitness_datasets_dir() / f"{dataset_id}.bak.json"
    _atomic_write_json(path, bak_path, f"fitness_dataset_{dataset_id}_", doc)


def _write_index(doc: dict[str, Any]) -> None:
    _atomic_write_json(_fitness_data_index_path(), _fitness_data_index_bak_path(), "fitness_datasets_index_", doc)


def _resolve_dataset(index: dict[str, Any], requested_dataset_id: str) -> tuple[str, dict[str, Any] | None]:
    dataset_id = requested_dataset_id or str(index.get("selectedDatasetId") or "")
    if not dataset_id:
        return "", None
    return dataset_id, _load_dataset_doc(dataset_id)


def _sheet_rows(ws: Any) -> list[list[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _extract_team_names_from_rows(rows: list[list[Any]]) -> list[str]:
    if len(rows) == 0:
        return []
    maxc = max((len(r) for r in rows), default=0)
    if maxc == 0:
        return []

    header = [str(_to_cell_value(rows[0][c]) if c < len(rows[0]) else "").strip() for c in range(maxc)]
    team_idx, team_col = _pick_team_column(header)
    names: list[str] = []

    if team_idx >= 0:
        for row in rows[1:]:
            value = str(_to_cell_value(row[team_idx] if team_idx < len(row) else "")).strip()
            if not value:
                continue
            if value not in names:
                names.append(value)
            if len(names) >= 2:
                break

    if len(names) >= 2:
        return names

    # fallback: search first two repeated non-empty strings that look like team names
    candidate_count: dict[str, int] = {}
    for row in rows:
        for cell in row:
            text = str(_to_cell_value(cell)).strip()
            if not text:
                continue
            if len(text) > 40:
                continue
            # ignore generic labels
            norm = _normalize_token(text)
            if norm in {"team", "player", "stat type", "total overall", "home", "away", "主队", "客队", "数据"}:
                continue
            candidate_count[text] = candidate_count.get(text, 0) + 1

    ordered = sorted(candidate_count.items(), key=lambda x: (-x[1], x[0]))
    for name, _ in ordered:
        names.append(name)
        if len(names) >= 2:
            break
    return names[:2]


def _looks_like_valid_team_name(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    lower = text.lower()
    if lower in {"#ref!", "home", "away", "主队", "客队", "数据", "team", "球队总览"}:
        return False
    if _to_float(text) is not None:
        return False
    return True


def _extract_team_names_from_named_overviews(named_sheets: list[tuple[str, list[list[Any]]]] | None) -> list[str]:
    if not named_sheets:
        return []
    names: list[str] = []
    for key in ("home team overview", "away team overview"):
        for sheet_name, rows in named_sheets:
            if key not in _normalize_header_name(sheet_name):
                continue
            maxc = max((len(r) for r in rows), default=0)
            for ridx in range(min(len(rows), 10)):
                row = rows[ridx]
                for cidx in range(min(maxc, 6)):
                    text = str(_to_cell_value(row[cidx] if cidx < len(row) else "")).strip()
                    if _looks_like_valid_team_name(text):
                        names.append(text)
                        break
                if len(names) >= (2 if key == "away team overview" else 1):
                    break
            break
    unique: list[str] = []
    for name in names:
        if name not in unique:
            unique.append(name)
    return unique[:2]


def _build_team_doc_from_standard(rows: list[list[Any]]) -> tuple[dict[str, Any] | None, str]:
    if len(rows) < 2:
        return (None, "")

    maxc = max((len(r) for r in rows), default=0)
    if maxc <= 0:
        return (None, "")

    headers = [str(_to_cell_value(rows[0][c]) if c < len(rows[0]) else "").strip() for c in range(maxc)]
    team_idx, team_col = _pick_team_column(headers)
    if team_idx < 0:
        return (None, "")

    used_ids: dict[str, int] = {}
    teams: list[dict[str, Any]] = []

    for row in rows[1:]:
        raw: dict[str, Any] = {}
        numeric_found = False
        for idx, col in enumerate(headers):
            value = _to_cell_value(row[idx] if idx < len(row) else None)
            raw[col] = value
            if idx != team_idx and _to_float(value) is not None:
                numeric_found = True

        team_name = str(raw.get(team_col, "")).strip()
        if not team_name or not numeric_found:
            continue

        teams.append(
            {
                "id": _make_id(team_name, used_ids, "team"),
                "team": team_name,
                "raw": raw,
            }
        )

        if len(teams) >= 2:
            break

    if len(teams) < 2:
        return (None, "")

    numeric_columns: list[str] = []
    for col in headers:
        if col == team_col:
            continue
        if all(_to_float(team.get("raw", {}).get(col)) is not None for team in teams):
            numeric_columns.append(col)

    if len(numeric_columns) == 0:
        return (None, "")

    return (
        {
            "teamColumn": team_col,
            "allColumns": headers,
            "numericColumns": numeric_columns,
            "teams": teams[:2],
        },
        "",
    )


def _find_overview_columns(headers_row: list[Any], next_row: list[Any] | None = None) -> tuple[int, int, int]:
    tokens = [_normalize_token(_to_cell_value(v)) for v in headers_row]
    home_label_idx = -1
    away_label_idx = -1
    metric_idx = -1

    for idx, token in enumerate(tokens):
        if token in {"home", "主队"} and home_label_idx < 0:
            home_label_idx = idx
        if token in {"away", "客队"} and away_label_idx < 0:
            away_label_idx = idx
        if token in {"data", "stat", "metric", "数据", "指标", "项目"} and metric_idx < 0:
            metric_idx = idx

    if home_label_idx < 0 or away_label_idx < 0 or metric_idx < 0:
        return (-1, -1, -1)

    # Many vendor templates use "label in one column, value in the next column".
    home_value_idx = home_label_idx
    away_value_idx = away_label_idx
    if next_row is not None:
        home_direct = _to_float(next_row[home_label_idx] if home_label_idx < len(next_row) else None)
        away_direct = _to_float(next_row[away_label_idx] if away_label_idx < len(next_row) else None)
        home_shift = _to_float(next_row[home_label_idx + 1] if home_label_idx + 1 < len(next_row) else None)
        away_shift = _to_float(next_row[away_label_idx + 1] if away_label_idx + 1 < len(next_row) else None)
        if home_direct is None and home_shift is not None:
            home_value_idx = home_label_idx + 1
        if away_direct is None and away_shift is not None:
            away_value_idx = away_label_idx + 1

    return (home_value_idx, metric_idx, away_value_idx)


def _build_team_doc_from_overview(rows: list[list[Any]], preferred_team_names: list[str]) -> tuple[dict[str, Any] | None, str]:
    if len(rows) < 3:
        return (None, "")

    home_idx = -1
    metric_idx = -1
    away_idx = -1
    header_row_index = -1

    for ridx, row in enumerate(rows[:10]):
        next_row = rows[ridx + 1] if ridx + 1 < len(rows) else None
        h, m, a = _find_overview_columns(row, next_row)
        if h >= 0:
            home_idx, metric_idx, away_idx = h, m, a
            header_row_index = ridx
            break

    if home_idx < 0:
        return (None, "")

    metrics: list[str] = []
    home_raw: dict[str, Any] = {}
    away_raw: dict[str, Any] = {}

    for row in rows[header_row_index + 1 :]:
        metric = str(_to_cell_value(row[metric_idx] if metric_idx < len(row) else "")).strip()
        if not metric:
            continue
        home_value = _to_float(row[home_idx] if home_idx < len(row) else None)
        away_value = _to_float(row[away_idx] if away_idx < len(row) else None)
        if home_value is None or away_value is None:
            continue
        metrics.append(metric)
        home_raw[metric] = home_value
        away_raw[metric] = away_value

    if len(metrics) < 3:
        return (None, "")

    team_a = preferred_team_names[0] if len(preferred_team_names) > 0 else "主队"
    team_b = preferred_team_names[1] if len(preferred_team_names) > 1 else "客队"

    home_raw_with_team = {"Team": team_a, **home_raw}
    away_raw_with_team = {"Team": team_b, **away_raw}

    teams = [
        {"id": _make_id(team_a, {}, "team"), "team": team_a, "raw": home_raw_with_team},
        {"id": _make_id(team_b, {team_a.lower(): 1}, "team"), "team": team_b, "raw": away_raw_with_team},
    ]

    return (
        {
            "teamColumn": "Team",
            "allColumns": ["Team", *metrics],
            "numericColumns": metrics,
            "teams": teams,
        },
        "",
    )


def _build_player_doc_from_standard(rows: list[list[Any]]) -> tuple[dict[str, Any] | None, str]:
    if len(rows) < 2:
        return (None, "")

    maxc = max((len(r) for r in rows), default=0)
    if maxc <= 0:
        return (None, "")

    raw_headers = [str(_to_cell_value(rows[0][c]) if c < len(rows[0]) else "").strip() for c in range(maxc)]
    headers = _normalize_headers(raw_headers)

    player_idx, _ = _pick_player_column(raw_headers)
    if player_idx < 0:
        player_idx, _ = _pick_player_column(headers)
    if player_idx < 0:
        player_idx, _ = _pick_player_column_by_values(rows, headers)
    if player_idx < 0:
        return (None, "第2个 sheet 未识别到球员列（可将姓名放在第2列，或使用 Player/Name/姓名 表头）")
    player_col = headers[player_idx]

    used_ids: dict[str, int] = {}
    players: list[dict[str, Any]] = []

    for row in rows[1:]:
        raw: dict[str, Any] = {}
        numeric_found = False
        for idx, col in enumerate(headers):
            value = _to_cell_value(row[idx] if idx < len(row) else None)
            raw[col] = value
            if idx != player_idx and _to_float(value) is not None:
                numeric_found = True

        player_name = str(raw.get(player_col, "")).strip()
        if not player_name or not numeric_found:
            continue

        players.append(
            {
                "id": _make_id(player_name, used_ids, "player"),
                "player": player_name,
                "raw": raw,
            }
        )

    if len(players) == 0:
        return (None, "")

    numeric_columns: list[str] = []
    for col in headers:
        if col == player_col:
            continue
        if any(_to_float(player.get("raw", {}).get(col)) is not None for player in players):
            numeric_columns.append(col)

    if len(numeric_columns) == 0:
        return (None, "")

    return (
        {
            "playerColumn": player_col,
            "allColumns": headers,
            "numericColumns": numeric_columns,
            "players": players,
        },
        "",
    )


def _build_player_doc_from_team_overview(rows: list[list[Any]]) -> tuple[dict[str, Any] | None, str]:
    if len(rows) < 6:
        return (None, "")

    maxc = max((len(r) for r in rows), default=0)
    if maxc <= 0:
        return (None, "")

    header_row_index = -1
    headers: list[str] = []
    for ridx in range(min(len(rows), 40)):
        candidate = [str(_to_cell_value(rows[ridx][c]) if c < len(rows[ridx]) else "").strip() for c in range(maxc)]
        joined = " ".join([v for v in candidate if v]).lower()
        if "出场时间" in joined and "总距离" in joined and "平均速度" in joined:
            header_row_index = ridx
            headers = candidate
            break

    if header_row_index < 0:
        return (None, "")

    data_rows = rows[header_row_index + 1 :]
    if len(data_rows) == 0:
        return (None, "")

    player_idx = -1
    best_player_score = -1
    for idx in range(min(maxc, 8)):
        non_empty = 0
        text_like = 0
        numbered = 0
        for row in data_rows[:120]:
            cell = _to_cell_value(row[idx] if idx < len(row) else None)
            text = str(cell).strip()
            if not text:
                continue
            non_empty += 1
            if _to_float(text) is None:
                text_like += 1
            if _looks_like_numbered_player_name(text):
                numbered += 1
        if non_empty < 3:
            continue
        score = text_like + numbered * 3
        if score > best_player_score:
            best_player_score = score
            player_idx = idx

    if player_idx < 0:
        return (None, "")

    metric_columns: list[tuple[int, str]] = []
    for idx, header in enumerate(headers):
        name = str(header or "").strip()
        if not name or idx == player_idx:
            continue
        numeric_hits = 0
        for row in data_rows[:160]:
            value = _to_float(row[idx] if idx < len(row) else None)
            if value is not None:
                numeric_hits += 1
        if numeric_hits >= 3:
            metric_columns.append((idx, name))

    if len(metric_columns) < 3:
        return (None, "")

    used_ids: dict[str, int] = {}
    players: list[dict[str, Any]] = []
    for row in data_rows:
        player_name = str(_to_cell_value(row[player_idx] if player_idx < len(row) else "")).strip()
        if not player_name:
            continue
        raw = {headers[player_idx] or "Player": player_name}
        numeric_found = False
        for col_idx, metric_name in metric_columns:
            value = _to_cell_value(row[col_idx] if col_idx < len(row) else None)
            raw[metric_name] = value
            if _to_float(value) is not None:
                numeric_found = True
        if not numeric_found:
            continue
        players.append(
            {
                "id": _make_id(player_name, used_ids, "player"),
                "player": player_name,
                "raw": raw,
            }
        )

    if len(players) == 0:
        return (None, "")

    numeric_columns = [name for _, name in metric_columns if any(_to_float(p.get("raw", {}).get(name)) is not None for p in players)]
    if len(numeric_columns) == 0:
        return (None, "")

    player_col_name = str(headers[player_idx] or "").strip() or "Player"
    all_columns = [player_col_name, *numeric_columns]
    normalized_players: list[dict[str, Any]] = []
    for player in players:
        raw = player.get("raw", {})
        normalized_raw = {player_col_name: raw.get(player_col_name, player.get("player", ""))}
        for col in numeric_columns:
            normalized_raw[col] = raw.get(col, "")
        normalized_players.append(
            {
                "id": player.get("id"),
                "player": player.get("player"),
                "raw": normalized_raw,
            }
        )

    return (
        {
            "playerColumn": player_col_name,
            "allColumns": all_columns,
            "numericColumns": numeric_columns,
            "players": normalized_players,
        },
        "",
    )


def _find_column(headers: list[str], targets: set[str]) -> list[int]:
    out: list[int] = []
    for idx, header in enumerate(headers):
        if _normalize_header_name(header) in targets:
            out.append(idx)
    return out


def _build_player_doc_from_stat_blocks(rows: list[list[Any]]) -> tuple[dict[str, Any] | None, str]:
    if len(rows) < 3:
        return (None, "")

    maxc = max((len(r) for r in rows), default=0)
    if maxc <= 0:
        return (None, "")

    # Try to find header row containing Player / Stat Type / Total Overall
    header_row_index = -1
    header: list[str] = []
    for ridx in range(min(6, len(rows))):
        candidate = [str(_to_cell_value(rows[ridx][c]) if c < len(rows[ridx]) else "").strip() for c in range(maxc)]
        if any(_normalize_header_name(v) == "player" for v in candidate) and any(_normalize_header_name(v) == "stat type" for v in candidate):
            header_row_index = ridx
            header = candidate
            break

    if header_row_index < 0:
        return (None, "")

    player_cols = _find_column(header, {"player"})
    stat_cols = _find_column(header, {"stat type"})
    total_cols = _find_column(header, {"total overall"})

    blocks: list[tuple[int, int, int]] = []
    for pcol in player_cols:
        # Vendor sheets may leave several spacer columns between Player and Stat Type.
        stat_candidates = [s for s in stat_cols if pcol < s <= pcol + 14]
        if not stat_candidates:
            continue
        scol = stat_candidates[0]
        total_candidates = [t for t in total_cols if scol < t <= scol + 6]
        if not total_candidates:
            continue
        tcol = total_candidates[0]
        blocks.append((pcol, scol, tcol))

    if not blocks:
        return (None, "")

    players_map: dict[str, dict[str, Any]] = {}

    for pcol, scol, tcol in blocks:
        current_player = ""
        for row in rows[header_row_index + 1 :]:
            player_cell = str(_to_cell_value(row[pcol] if pcol < len(row) else "")).strip()
            if player_cell:
                current_player = player_cell
            if not current_player:
                continue

            stat_code = str(_to_cell_value(row[scol] if scol < len(row) else "")).strip()
            if not stat_code:
                continue

            value = _to_float(row[tcol] if tcol < len(row) else None)
            if value is None:
                continue

            metric = f"Stat {stat_code}"
            players_map.setdefault(current_player, {})[metric] = value

    if len(players_map) == 0:
        return (None, "")

    metric_set: set[str] = set()
    for pdata in players_map.values():
        metric_set.update(pdata.keys())
    numeric_columns = sorted(metric_set, key=lambda x: (int(x.split(" ", 1)[1]) if x.split(" ", 1)[1].isdigit() else 9999, x))

    if len(numeric_columns) == 0:
        return (None, "")

    used_ids: dict[str, int] = {}
    players: list[dict[str, Any]] = []
    for player_name, metric_map in players_map.items():
        raw = {"Player": player_name}
        for metric in numeric_columns:
            raw[metric] = metric_map.get(metric, "")
        players.append(
            {
                "id": _make_id(player_name, used_ids, "player"),
                "player": player_name,
                "raw": raw,
            }
        )

    return (
        {
            "playerColumn": "Player",
            "allColumns": ["Player", *numeric_columns],
            "numericColumns": numeric_columns,
            "players": players,
        },
        "",
    )


def _build_team_doc(
    sheet_rows_1: list[list[Any]],
    sheet_rows_2: list[list[Any]],
    named_sheets: list[tuple[str, list[list[Any]]]] | None = None,
) -> tuple[dict[str, Any] | None, str]:
    # Prefer the sheet explicitly named "Overview" for team radar data.
    if named_sheets:
        named_overview_team_names = _extract_team_names_from_named_overviews(named_sheets)
        for sheet_name, rows in named_sheets:
            if _normalize_header_name(sheet_name) == "overview":
                team_names = named_overview_team_names or _extract_team_names_from_rows(rows)
                doc, _ = _build_team_doc_from_overview(rows, team_names)
                if doc is not None:
                    return (doc, "")
                doc, _ = _build_team_doc_from_standard(rows)
                if doc is not None:
                    return (doc, "")

    team_names = _extract_team_names_from_named_overviews(named_sheets)
    if len(team_names) < 2:
        team_names = _extract_team_names_from_rows(sheet_rows_1)
    if len(team_names) < 2:
        team_names = _extract_team_names_from_rows(sheet_rows_2)

    parsers = [
        lambda: _build_team_doc_from_standard(sheet_rows_1),
        lambda: _build_team_doc_from_standard(sheet_rows_2),
        lambda: _build_team_doc_from_overview(sheet_rows_1, team_names),
        lambda: _build_team_doc_from_overview(sheet_rows_2, team_names),
    ]

    if named_sheets:
        for _, rows in named_sheets:
            parsers.append(lambda rows=rows: _build_team_doc_from_standard(rows))
            parsers.append(lambda rows=rows: _build_team_doc_from_overview(rows, team_names))

    for parser in parsers:
        doc, _ = parser()
        if doc is not None:
            return (doc, "")

    return (None, "未识别到两队有效体能数据（优先读取 Overview sheet）")


def _build_player_doc(
    sheet_rows_1: list[list[Any]],
    sheet_rows_2: list[list[Any]],
    named_sheets: list[tuple[str, list[list[Any]]]] | None = None,
    preferred_side: str = "home",
) -> tuple[dict[str, Any] | None, str, dict[str, Any]]:
    preferred_side = _normalize_player_overview_side(preferred_side)
    meta = {
        "requestedSide": preferred_side,
        "usedSide": "",
        "usedSheetName": "",
        "fallbackUsed": False,
    }

    if named_sheets:
        preferred_tokens = [
            ("home", "home team overview"),
            ("away", "away team overview"),
        ]
        if preferred_side == "away":
            preferred_tokens = [("away", "away team overview"), ("home", "home team overview")]

        for order_idx, (side, token) in enumerate(preferred_tokens):
            for sheet_name, rows in named_sheets:
                if token in _normalize_header_name(sheet_name):
                    doc, _ = _build_player_doc_from_team_overview(rows)
                    if doc is not None:
                        meta["usedSide"] = side
                        meta["usedSheetName"] = sheet_name
                        meta["fallbackUsed"] = order_idx > 0
                        return (doc, "", meta)

    candidates: list[dict[str, Any]] = []
    parse_errors: list[str] = []

    parser_rows: list[list[list[Any]]] = [sheet_rows_2, sheet_rows_1]
    if named_sheets:
        known = {id(sheet_rows_1), id(sheet_rows_2)}
        for _, rows in named_sheets:
            if id(rows) not in known:
                parser_rows.append(rows)

    for rows in parser_rows:
        for parser in (
            _build_player_doc_from_team_overview,
            _build_player_doc_from_standard,
            _build_player_doc_from_stat_blocks,
        ):
            doc, err = parser(rows)
            if doc is not None:
                candidates.append(doc)
            elif err:
                parse_errors.append(err)

    if candidates:
        best = max(candidates, key=_score_player_doc)
        if not meta["usedSide"]:
            meta["usedSide"] = preferred_side
        return (best, "", meta)

    if len(parse_errors) > 0:
        return (None, parse_errors[0], meta)
    return (None, "第1/2个 sheet 未识别到有效球员体能数据", meta)


@fitness_data_bp.route("/api/fitness-data/import-excel", methods=["POST"])
def import_fitness_excel():
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "only .xlsx is supported"}), 400

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"invalid excel file: {exc}"}), 400

    sheets = wb.worksheets
    if len(sheets) < 2:
        wb.close()
        return jsonify({"ok": False, "error": "excel 至少需要两个 sheet"}), 400

    sheet_1 = sheets[0]
    sheet_2 = sheets[1]
    rows_1 = _sheet_rows(sheet_1)
    rows_2 = _sheet_rows(sheet_2)
    requested_player_overview_side = _normalize_player_overview_side(request.form.get("playerOverviewSide"))
    named_sheet_rows = [(ws.title, _sheet_rows(ws)) for ws in sheets]
    team_doc, team_err = _build_team_doc(rows_1, rows_2, named_sheet_rows)
    if team_doc is None:
        wb.close()
        return jsonify({"ok": False, "error": team_err}), 400

    player_doc, player_err, player_doc_meta = _build_player_doc(rows_1, rows_2, named_sheet_rows, requested_player_overview_side)
    if player_doc is None:
        wb.close()
        return jsonify({"ok": False, "error": player_err}), 400

    wb.close()

    updated_at = _iso_now()
    dataset_id = f"fds_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"
    doc = {
        "version": VERSION,
        "datasetId": dataset_id,
        "updatedAt": updated_at,
        "source": {
            "filename": file.filename,
            "sheet1Name": sheet_1.title,
            "sheet2Name": sheet_2.title,
            "playerOverviewSideRequested": player_doc_meta.get("requestedSide", requested_player_overview_side),
            "playerOverviewSideUsed": player_doc_meta.get("usedSide", ""),
            "playerOverviewSheetUsed": player_doc_meta.get("usedSheetName", ""),
            "playerOverviewFallbackUsed": bool(player_doc_meta.get("fallbackUsed")),
        },
        "teamSheet": team_doc,
        "playerSheet": player_doc,
    }

    try:
        _write_dataset_doc(dataset_id, doc)
        index = _load_index()
        datasets = [d for d in index.get("datasets", []) if isinstance(d, dict)]
        datasets.insert(
            0,
            {
                "id": dataset_id,
                "name": f"{file.filename} ({updated_at[:19]})",
                "updatedAt": updated_at,
                "teamCount": len(team_doc.get("teams", [])),
                "teamMetricCount": len(team_doc.get("numericColumns", [])),
                "playerCount": len(player_doc.get("players", [])),
                "playerMetricCount": len(player_doc.get("numericColumns", [])),
                "sourceFile": file.filename,
            },
        )
        index["datasets"] = datasets
        index["selectedDatasetId"] = dataset_id
        index["updatedAt"] = updated_at
        _write_index(index)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"write failed: {exc}"}), 500

    return jsonify(
        {
            "ok": True,
            "datasetId": dataset_id,
            "updatedAt": updated_at,
            "teamCount": len(team_doc.get("teams", [])),
            "teamMetricCount": len(team_doc.get("numericColumns", [])),
            "playerCount": len(player_doc.get("players", [])),
            "playerMetricCount": len(player_doc.get("numericColumns", [])),
            "playerOverviewSideRequested": player_doc_meta.get("requestedSide", requested_player_overview_side),
            "playerOverviewSideUsed": player_doc_meta.get("usedSide", ""),
            "playerOverviewSheetUsed": player_doc_meta.get("usedSheetName", ""),
            "playerOverviewFallbackUsed": bool(player_doc_meta.get("fallbackUsed")),
        }
    )


@fitness_data_bp.route("/api/fitness-data/datasets", methods=["GET"])
def get_fitness_datasets():
    try:
        index = _load_index()
        return jsonify(
            {
                "ok": True,
                "datasets": index.get("datasets", []),
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "updatedAt": index.get("updatedAt"),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


@fitness_data_bp.route("/api/fitness-data", methods=["GET"])
def get_fitness_dataset():
    try:
        index = _load_index()
        dataset_id, doc = _resolve_dataset(index, str(request.args.get("datasetId") or ""))
        if doc is None:
            return jsonify(
                {
                    "ok": True,
                    "datasetId": dataset_id,
                    "selectedDatasetId": index.get("selectedDatasetId", ""),
                    "updatedAt": None,
                    "data": None,
                }
            )

        return jsonify(
            {
                "ok": True,
                "datasetId": dataset_id,
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "updatedAt": doc.get("updatedAt"),
                "data": doc,
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


@fitness_data_bp.route("/api/fitness-data/datasets/<dataset_id>", methods=["DELETE"])
def delete_fitness_dataset(dataset_id: str):
    try:
        index = _load_index()
        datasets = [d for d in index.get("datasets", []) if isinstance(d, dict)]
        matched = next((d for d in datasets if d.get("id") == dataset_id), None)
        if matched is None:
            return jsonify({"ok": False, "error": "dataset not found"}), 404

        path = _dataset_file_path(dataset_id)
        if path.exists():
            path.unlink()

        remaining = [d for d in datasets if d.get("id") != dataset_id]
        index["datasets"] = remaining
        if index.get("selectedDatasetId") == dataset_id:
            index["selectedDatasetId"] = remaining[0].get("id") if remaining else ""
        index["updatedAt"] = _iso_now()
        _write_index(index)

        return jsonify(
            {
                "ok": True,
                "deletedDatasetId": dataset_id,
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "datasets": remaining,
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"delete failed: {exc}"}), 500
