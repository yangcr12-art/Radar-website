from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from pathlib import Path
from tempfile import NamedTemporaryFile
from threading import Lock
from typing import Any
from uuid import uuid4

from flask import Blueprint, jsonify, request
from openpyxl import load_workbook


VERSION = 1
WRITE_LOCK = Lock()
APP_DIR = Path(__file__).resolve().parents[2]
DATA_DIR = APP_DIR / "data"
CSL_STANDINGS_DATASETS_DIR = DATA_DIR / "csl_standings_datasets"
CSL_STANDINGS_INDEX_PATH = DATA_DIR / "csl_standings_datasets_index.json"
CSL_STANDINGS_INDEX_BAK_PATH = DATA_DIR / "csl_standings_datasets_index.json.bak"

csl_standings_bp = Blueprint("csl_standings_api", __name__)

LEAGUE_TOKEN_CSL = "中超"
STATUS_FINISHED_TOKENS = {"比赛结束", "完场", "finished", "ft"}
DEDUCTION_SEASON = "2026"
TEAM_DEDUCTION_POINTS_2026 = {
    "上海申花": 10,
    "天津津门虎": 10,
    "青岛海牛": 7,
    "山东泰山": 6,
    "河南": 6,
    "上海海港": 5,
    "北京国安": 5,
    "浙江": 5,
    "武汉三镇": 5,
}
TEAM_DEDUCTION_ALIASES = {
    "河南队": "河南",
    "河南足球俱乐部": "河南",
    "浙江职业足球俱乐部": "浙江",
    "浙江队": "浙江",
}

_HEADER_ALIASES = {
    "season": ["赛季", "season"],
    "league": ["联赛名称", "联赛", "league", "competition"],
    "round": ["轮次", "轮", "round", "matchweek"],
    "matchTime": ["比赛时间", "开球时间", "matchtime", "kickoff", "date", "time"],
    "homeTeam": ["主队俱乐部名称", "主队", "hometeam", "home"],
    "homeScore": ["主场队伍总得分", "主队得分", "主场进球", "homescore", "homegoals"],
    "awayTeam": ["客队俱乐部名称", "客队", "awayteam", "away"],
    "awayScore": ["客场队伍总得分", "客队得分", "客场进球", "awayscore", "awaygoals"],
    "status": ["比赛状态", "status", "matchstatus"],
}


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    CSL_STANDINGS_DATASETS_DIR.mkdir(parents=True, exist_ok=True)


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_token(value: Any) -> str:
    return "".join(str(value or "").strip().lower().split())


def _to_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    try:
        num = float(text)
    except ValueError:
        return None
    if not num.is_integer():
        return None
    return int(num)


def _to_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def _default_index() -> dict[str, Any]:
    return {
        "version": VERSION,
        "updatedAt": None,
        "selectedDatasetId": "",
        "datasets": [],
    }


def _dataset_file_path(dataset_id: str) -> Path:
    return CSL_STANDINGS_DATASETS_DIR / f"{dataset_id}.json"


def _atomic_write_json(path: Path, bak_path: Path, prefix: str, doc: dict[str, Any]) -> None:
    _ensure_data_dir()
    data = json.dumps(doc, ensure_ascii=False, indent=2)
    with WRITE_LOCK:
        if path.exists():
            bak_path.write_text(path.read_text(encoding="utf-8"), encoding="utf-8")
        with NamedTemporaryFile("w", encoding="utf-8", delete=False, dir=path.parent, prefix=prefix, suffix=".tmp") as tf:
            tf.write(data)
            tmp_path = Path(tf.name)
        os.replace(tmp_path, path)


def _load_index() -> dict[str, Any]:
    if CSL_STANDINGS_INDEX_PATH.exists():
        with CSL_STANDINGS_INDEX_PATH.open("r", encoding="utf-8") as f:
            idx = json.load(f)
            if isinstance(idx, dict):
                return {
                    "version": int(idx.get("version", VERSION)),
                    "updatedAt": idx.get("updatedAt"),
                    "selectedDatasetId": str(idx.get("selectedDatasetId") or ""),
                    "datasets": idx.get("datasets") if isinstance(idx.get("datasets"), list) else [],
                }
    return _default_index()


def _load_dataset_doc(dataset_id: str) -> dict[str, Any] | None:
    path = _dataset_file_path(dataset_id)
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _write_dataset_doc(dataset_id: str, doc: dict[str, Any]) -> None:
    path = _dataset_file_path(dataset_id)
    bak_path = CSL_STANDINGS_DATASETS_DIR / f"{dataset_id}.bak.json"
    _atomic_write_json(path, bak_path, f"csl_standings_dataset_{dataset_id}_", doc)


def _write_index(doc: dict[str, Any]) -> None:
    _atomic_write_json(CSL_STANDINGS_INDEX_PATH, CSL_STANDINGS_INDEX_BAK_PATH, "csl_standings_index_", doc)


def _resolve_dataset(index: dict[str, Any], requested_dataset_id: str) -> tuple[str, dict[str, Any] | None]:
    dataset_id = requested_dataset_id or str(index.get("selectedDatasetId") or "")
    if not dataset_id:
        return "", None
    return dataset_id, _load_dataset_doc(dataset_id)


def _resolve_header_indices(header_row: list[Any]) -> tuple[dict[str, int], list[str]]:
    header_values = [_normalize_token(_to_cell_value(cell)) for cell in header_row]
    found: dict[str, int] = {}
    for field, aliases in _HEADER_ALIASES.items():
        alias_tokens = {_normalize_token(alias) for alias in aliases}
        idx = next((i for i, value in enumerate(header_values) if value in alias_tokens), -1)
        if idx >= 0:
            found[field] = idx

    required_fields = ["season", "league", "round", "homeTeam", "homeScore", "awayTeam", "awayScore", "status"]
    missing = [field for field in required_fields if field not in found]
    return found, missing


def _is_finished_status(status: str) -> bool:
    token = _normalize_token(status)
    return token in {_normalize_token(item) for item in STATUS_FINISHED_TOKENS}


def _is_csl_league(league: str) -> bool:
    return LEAGUE_TOKEN_CSL in _normalize_text(league)


def _sort_seasons(seasons: set[str]) -> list[str]:
    def _key(value: str) -> tuple[int, str]:
        token = str(value or "").strip()
        if token.isdigit():
            return (0, f"{int(token):09d}")
        return (1, token)

    return sorted([s for s in seasons if s], key=_key, reverse=True)


def _normalize_team_for_deduction(team: str) -> str:
    normalized = _normalize_text(team)
    return TEAM_DEDUCTION_ALIASES.get(normalized, normalized)


def _deduction_for_team(season: str, team: str) -> int:
    if str(season or "") != DEDUCTION_SEASON:
        return 0
    key = _normalize_team_for_deduction(team)
    return int(TEAM_DEDUCTION_POINTS_2026.get(key, 0))


def _assign_rank(rows: list[dict[str, Any]], points_key: str, rank_key: str) -> list[dict[str, Any]]:
    sorted_rows = sorted(
        rows,
        key=lambda item: (
            -int(item.get(points_key, 0)),
            -int(item.get("goalDiff", 0)),
            -int(item.get("goalsFor", 0)),
            str(item.get("team", "")),
        ),
    )
    prev_tuple: tuple[int, int, int] | None = None
    current_rank = 0
    for idx, row in enumerate(sorted_rows, start=1):
        rank_tuple = (
            int(row.get(points_key, 0)),
            int(row.get("goalDiff", 0)),
            int(row.get("goalsFor", 0)),
        )
        if rank_tuple != prev_tuple:
            current_rank = idx
            prev_tuple = rank_tuple
        row[rank_key] = current_rank
    return sorted_rows


def _build_trend_payload(doc: dict[str, Any], season: str) -> dict[str, Any]:
    all_matches = doc.get("matches") if isinstance(doc.get("matches"), list) else []
    season_matches = [m for m in all_matches if str(m.get("season") or "") == str(season)]
    if not season_matches:
        return {
            "selectedSeason": season,
            "rounds": [],
            "teams": [],
            "standingsByRound": [],
            "trendSeriesByTeam": {},
        }

    rounds = sorted({int(m["round"]) for m in season_matches if isinstance(m.get("round"), int)})
    teams = sorted(
        {
            str(m.get("homeTeam") or "").strip()
            for m in season_matches
            if str(m.get("homeTeam") or "").strip()
        }
        | {
            str(m.get("awayTeam") or "").strip()
            for m in season_matches
            if str(m.get("awayTeam") or "").strip()
        }
    )

    stats_by_team: dict[str, dict[str, Any]] = {
        team: {
            "played": 0,
            "won": 0,
            "draw": 0,
            "lost": 0,
            "points": 0,
            "goalsFor": 0,
            "goalsAgainst": 0,
        }
        for team in teams
    }

    matches_by_round: dict[int, list[dict[str, Any]]] = {}
    for item in season_matches:
        round_num = int(item.get("round"))
        matches_by_round.setdefault(round_num, []).append(item)

    standings_by_round: list[dict[str, Any]] = []
    trend_series_by_team: dict[str, list[dict[str, Any]]] = {team: [] for team in teams}

    for round_num in rounds:
        for match in matches_by_round.get(round_num, []):
            home_team = str(match.get("homeTeam") or "").strip()
            away_team = str(match.get("awayTeam") or "").strip()
            home_score = int(match.get("homeScore"))
            away_score = int(match.get("awayScore"))
            if not home_team or not away_team:
                continue
            home = stats_by_team.setdefault(
                home_team,
                {"played": 0, "won": 0, "draw": 0, "lost": 0, "points": 0, "goalsFor": 0, "goalsAgainst": 0},
            )
            away = stats_by_team.setdefault(
                away_team,
                {"played": 0, "won": 0, "draw": 0, "lost": 0, "points": 0, "goalsFor": 0, "goalsAgainst": 0},
            )

            home["played"] += 1
            away["played"] += 1
            home["goalsFor"] += home_score
            home["goalsAgainst"] += away_score
            away["goalsFor"] += away_score
            away["goalsAgainst"] += home_score

            if home_score > away_score:
                home["won"] += 1
                home["points"] += 3
                away["lost"] += 1
            elif home_score < away_score:
                away["won"] += 1
                away["points"] += 3
                home["lost"] += 1
            else:
                home["draw"] += 1
                away["draw"] += 1
                home["points"] += 1
                away["points"] += 1

        rows: list[dict[str, Any]] = []
        for team, stat in stats_by_team.items():
            goals_for = int(stat["goalsFor"])
            goals_against = int(stat["goalsAgainst"])
            points_raw = int(stat["points"])
            deduction = _deduction_for_team(season, team)
            rows.append(
                {
                    "team": team,
                    "played": int(stat["played"]),
                    "won": int(stat["won"]),
                    "draw": int(stat["draw"]),
                    "lost": int(stat["lost"]),
                    "deduction": int(deduction),
                    "pointsRaw": points_raw,
                    "pointsNet": points_raw - int(deduction),
                    "goalsFor": goals_for,
                    "goalsAgainst": goals_against,
                    "goalDiff": goals_for - goals_against,
                }
            )

        _assign_rank(rows, "pointsRaw", "rankRaw")
        ranked_rows = _assign_rank(rows, "pointsNet", "rankNet")
        for row in ranked_rows:
            row["points"] = int(row.get("pointsRaw", 0))
            row["rank"] = int(row.get("rankRaw", 0))
        standings_by_round.append({"round": round_num, "rows": ranked_rows})

        row_by_team = {str(row["team"]): row for row in ranked_rows}
        for team in teams:
            row = row_by_team.get(team)
            if row is None:
                continue
            trend_series_by_team[team].append(
                {
                    "round": round_num,
                    "deduction": int(row.get("deduction", 0)),
                    "pointsRaw": int(row.get("pointsRaw", 0)),
                    "pointsNet": int(row.get("pointsNet", 0)),
                    "rankRaw": int(row.get("rankRaw", 0)),
                    "rankNet": int(row.get("rankNet", 0)),
                    "points": int(row.get("pointsRaw", 0)),
                    "rank": int(row.get("rankRaw", 0)),
                    "goalsFor": int(row["goalsFor"]),
                    "goalsAgainst": int(row["goalsAgainst"]),
                }
            )

    return {
        "selectedSeason": season,
        "rounds": rounds,
        "teams": teams,
        "standingsByRound": standings_by_round,
        "trendSeriesByTeam": trend_series_by_team,
    }


@csl_standings_bp.route("/api/csl-standings/import-excel", methods=["POST"])
def import_csl_standings_excel():
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "only .xlsx is supported"}), 400

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"invalid excel file: {exc}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return jsonify({"ok": False, "error": "excel must contain header and data rows"}), 400

    header = list(rows[0])
    field_indices, missing = _resolve_header_indices(header)
    if missing:
        return jsonify({"ok": False, "error": f"missing required columns: {', '.join(missing)}"}), 400

    imported_matches: list[dict[str, Any]] = []
    seasons: set[str] = set()
    teams: set[str] = set()
    invalid_row_count = 0
    deduped_row_count = 0
    seen_match_keys: set[str] = set()

    for raw_row in rows[1:]:
        cells = list(raw_row)

        def _value(field: str) -> Any:
            idx = field_indices.get(field, -1)
            return cells[idx] if idx >= 0 and idx < len(cells) else None

        season = _normalize_text(_to_cell_value(_value("season")))
        league = _normalize_text(_to_cell_value(_value("league")))
        round_num = _to_int(_to_cell_value(_value("round")))
        match_time = _normalize_text(_to_cell_value(_value("matchTime")))
        home_team = _normalize_text(_to_cell_value(_value("homeTeam")))
        away_team = _normalize_text(_to_cell_value(_value("awayTeam")))
        home_score = _to_int(_to_cell_value(_value("homeScore")))
        away_score = _to_int(_to_cell_value(_value("awayScore")))
        status = _normalize_text(_to_cell_value(_value("status")))

        if not _is_csl_league(league):
            continue
        if not _is_finished_status(status):
            continue

        if not season or round_num is None or not home_team or not away_team or home_score is None or away_score is None:
            invalid_row_count += 1
            continue

        dedupe_key = "::".join(
            [
                season,
                str(round_num),
                match_time,
                home_team,
                away_team,
                str(home_score),
                str(away_score),
            ]
        )
        if dedupe_key in seen_match_keys:
            deduped_row_count += 1
            continue
        seen_match_keys.add(dedupe_key)

        imported_matches.append(
            {
                "season": season,
                "league": league,
                "round": int(round_num),
                "matchTime": match_time,
                "homeTeam": home_team,
                "awayTeam": away_team,
                "homeScore": int(home_score),
                "awayScore": int(away_score),
                "status": status,
            }
        )
        seasons.add(season)
        teams.add(home_team)
        teams.add(away_team)

    if not imported_matches:
        return jsonify({"ok": False, "error": "no valid CSL finished matches found"}), 400

    imported_matches.sort(
        key=lambda item: (
            str(item.get("season") or ""),
            int(item.get("round") or 0),
            str(item.get("matchTime") or ""),
            str(item.get("homeTeam") or ""),
            str(item.get("awayTeam") or ""),
        )
    )

    updated_at = _iso_now()
    dataset_id = f"csd_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"
    seasons_sorted = _sort_seasons(seasons)

    doc_to_save = {
        "version": VERSION,
        "datasetId": dataset_id,
        "updatedAt": updated_at,
        "source": {
            "filename": file.filename,
            "league": LEAGUE_TOKEN_CSL,
            "importedAt": updated_at,
        },
        "seasons": seasons_sorted,
        "matches": imported_matches,
        "stats": {
            "matchCount": len(imported_matches),
            "teamCount": len(teams),
            "invalidRowCount": invalid_row_count,
            "dedupedRowCount": deduped_row_count,
        },
    }

    try:
        _write_dataset_doc(dataset_id, doc_to_save)
        index = _load_index()
        datasets = [d for d in index.get("datasets", []) if isinstance(d, dict)]
        datasets.insert(
            0,
            {
                "id": dataset_id,
                "name": f"{file.filename} ({updated_at[:19]})",
                "updatedAt": updated_at,
                "sourceFile": file.filename,
                "seasonCount": len(seasons_sorted),
                "matchCount": len(imported_matches),
                "teamCount": len(teams),
            },
        )
        index["datasets"] = datasets
        index["selectedDatasetId"] = dataset_id
        index["updatedAt"] = updated_at
        _write_index(index)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"write failed: {exc}"}), 500

    return jsonify(
        {
            "ok": True,
            "datasetId": dataset_id,
            "updatedAt": updated_at,
            "seasons": seasons_sorted,
            "matchCount": len(imported_matches),
            "teamCount": len(teams),
            "invalidRowCount": invalid_row_count,
            "dedupedRowCount": deduped_row_count,
        }
    )


@csl_standings_bp.route("/api/csl-standings/datasets", methods=["GET"])
def get_csl_standings_datasets():
    try:
        index = _load_index()
        return jsonify(
            {
                "ok": True,
                "datasets": index.get("datasets", []),
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "updatedAt": index.get("updatedAt"),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


@csl_standings_bp.route("/api/csl-standings", methods=["GET"])
def get_csl_standings_dataset():
    try:
        index = _load_index()
        dataset_id, doc = _resolve_dataset(index, str(request.args.get("datasetId") or ""))
        if doc is None:
            return jsonify(
                {
                    "ok": True,
                    "datasetId": dataset_id,
                    "selectedDatasetId": index.get("selectedDatasetId", ""),
                    "updatedAt": None,
                    "data": None,
                }
            )

        seasons = doc.get("seasons") if isinstance(doc.get("seasons"), list) else []
        seasons = [str(s) for s in seasons if str(s)]
        requested_season = str(request.args.get("season") or "")
        if requested_season and requested_season in seasons:
            season = requested_season
        else:
            season = seasons[0] if seasons else ""

        trend_payload = _build_trend_payload(doc, season)
        data = {
            "datasetId": str(doc.get("datasetId") or dataset_id),
            "updatedAt": doc.get("updatedAt"),
            "source": doc.get("source") if isinstance(doc.get("source"), dict) else {},
            "stats": doc.get("stats") if isinstance(doc.get("stats"), dict) else {},
            "seasons": seasons,
            **trend_payload,
        }

        return jsonify(
            {
                "ok": True,
                "datasetId": dataset_id,
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "updatedAt": doc.get("updatedAt"),
                "data": data,
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


@csl_standings_bp.route("/api/csl-standings/datasets/<dataset_id>", methods=["DELETE"])
def delete_csl_standings_dataset(dataset_id: str):
    try:
        index = _load_index()
        datasets = [d for d in index.get("datasets", []) if isinstance(d, dict)]
        matched = next((d for d in datasets if d.get("id") == dataset_id), None)
        if matched is None:
            return jsonify({"ok": False, "error": "dataset not found"}), 404

        path = _dataset_file_path(dataset_id)
        if path.exists():
            path.unlink()

        remaining = [d for d in datasets if d.get("id") != dataset_id]
        index["datasets"] = remaining
        if index.get("selectedDatasetId") == dataset_id:
            index["selectedDatasetId"] = remaining[0].get("id") if remaining else ""
        index["updatedAt"] = _iso_now()
        _write_index(index)

        return jsonify(
            {
                "ok": True,
                "deletedDatasetId": dataset_id,
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "datasets": remaining,
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"delete failed: {exc}"}), 500
