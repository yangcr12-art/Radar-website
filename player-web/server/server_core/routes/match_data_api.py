from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from pathlib import Path
from tempfile import NamedTemporaryFile
from threading import Lock
from typing import Any
from uuid import uuid4

from flask import Blueprint, jsonify, request
from openpyxl import load_workbook
from server_core.services.auth_config import get_primary_login_username
from server_core.services.ranking_service import (
    compute_player_metrics,
    is_lower_better_column,
)
from server_core.services.session_auth import get_authenticated_username
from server_core.services.user_storage import ensure_user_data_dir, user_data_file, user_data_subdir


VERSION = 1
WRITE_LOCK = Lock()
match_data_bp = Blueprint("match_data_api", __name__)


def _resolve_username() -> str:
    return get_authenticated_username(get_primary_login_username())


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _ensure_data_dir() -> None:
    ensure_user_data_dir(_resolve_username())
    user_data_subdir(_resolve_username(), "match_datasets")


def _match_datasets_dir() -> Path:
    return user_data_subdir(_resolve_username(), "match_datasets")


def _match_data_index_path() -> Path:
    return user_data_file(_resolve_username(), "match_datasets_index.json")


def _match_data_index_bak_path() -> Path:
    return user_data_file(_resolve_username(), "match_datasets_index.json.bak")


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = text.replace(",", "")
        try:
            return float(text)
        except ValueError:
            return None
    return None


def _to_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def _count_numeric_cells(row: list[Any], maxc: int) -> int:
    count = 0
    for c in range(maxc):
        value = row[c] if c < len(row) else None
        if _to_float(value) is not None:
            count += 1
    return count


def _split_tokens(text: str) -> list[str]:
    base = str(text or "").replace("／", "/").strip()
    if not base:
        return []
    return [item.strip() for item in base.split("/") if item.strip()]


def _split_special_cases(en_anchor: str, zh_anchor: str, span: int) -> list[str]:
    en_norm = str(en_anchor or "").lower().replace(" ", "")
    zh_norm = str(zh_anchor or "").replace(" ", "")
    if span == 3 and ("shots/ontarget" in en_norm or "射门/射正" in zh_norm):
        return ["Shots", "Shots on target", "Shots on target, %"]
    if span == 4 and ("recoveries/low/medium/high" in en_norm or "夺回球权/低位/中位/高位" in zh_norm):
        return ["Recoveries", "Recoveries - Low", "Recoveries - Medium", "Recoveries - High"]
    if span == 4 and ("losses/low/medium/high" in en_norm or "丢失球权/低位/中位/高位" in zh_norm):
        return ["Losses", "Losses - Low", "Losses - Medium", "Losses - High"]
    if span == 3 and ("penaltyareaentries" in en_norm or "攻入禁区" in zh_norm):
        return ["Penalty area entries", "Penalty area entries - Runs", "Penalty area entries - Crosses"]
    return []


def _build_span_header_names(en_anchor: str, zh_anchor: str, span: int) -> list[str]:
    if span <= 1:
        en = str(en_anchor or "").strip()
        return [en]

    special = _split_special_cases(en_anchor, zh_anchor, span)
    if special:
        return special

    en_tokens = _split_tokens(str(en_anchor or ""))
    if len(en_tokens) == span and span > 0:
        return en_tokens

    if span == 3 and len(en_tokens) == 2:
        base_en, qual_en = en_tokens
        return [base_en, f"{base_en}, {qual_en}", f"{base_en}, {qual_en} %"]

    if span == len(en_tokens) + 1 and len(en_tokens) >= 1:
        return en_tokens[:-1] + [en_tokens[-1], f"{en_tokens[-1]}, %"]

    en = str(en_anchor or "").strip() or "Metric"
    return [f"{en} #{idx + 1}" for idx in range(span)]


def _build_split_headers(row1: list[Any], row2: list[Any], maxc: int) -> list[str]:
    en_row = [str(_to_cell_value(row1[c]) if c < len(row1) else "").strip() for c in range(maxc)]
    zh_row = [str(_to_cell_value(row2[c]) if c < len(row2) else "").strip() for c in range(maxc)]

    anchors = [idx for idx, value in enumerate(en_row) if value]
    if not anchors:
        return []

    headers = ["" for _ in range(maxc)]
    for i, anchor_idx in enumerate(anchors):
        next_anchor = anchors[i + 1] if i + 1 < len(anchors) else maxc
        span = max(1, next_anchor - anchor_idx)
        names = _build_span_header_names(en_row[anchor_idx], zh_row[anchor_idx], span)
        if len(names) != span:
            anchor_text = en_row[anchor_idx] or "Metric"
            names = (names + [f"{anchor_text} #{n + 1}" for n in range(span)])[:span]
        for offset in range(span):
            headers[anchor_idx + offset] = str(names[offset] or "").strip()

    # Deduplicate headers while keeping stable display names.
    seen: dict[str, int] = {}
    for idx, header in enumerate(headers):
        key = header.lower().strip()
        if not key:
            headers[idx] = f"Column {idx + 1}"
            key = headers[idx].lower()
        count = seen.get(key, 0) + 1
        seen[key] = count
        if count > 1:
            headers[idx] = f"{header} ({count})"
    return headers


def _make_id(name: str, used_ids: dict[str, int]) -> str:
    base = "".join(ch.lower() if ch.isalnum() else "_" for ch in name).strip("_")
    if not base:
        base = "team"
    if base not in used_ids:
        used_ids[base] = 1
        return base
    used_ids[base] += 1
    return f"{base}_{used_ids[base]}"


def _default_index() -> dict[str, Any]:
    return {
        "version": VERSION,
        "updatedAt": None,
        "selectedDatasetId": "",
        "datasets": [],
    }


def _dataset_file_path(dataset_id: str) -> Path:
    return _match_datasets_dir() / f"{dataset_id}.json"


def _atomic_write_json(path: Path, bak_path: Path, prefix: str, doc: dict[str, Any]) -> None:
    _ensure_data_dir()
    data = json.dumps(doc, ensure_ascii=False, indent=2)
    with WRITE_LOCK:
        if path.exists():
            bak_path.write_text(path.read_text(encoding="utf-8"), encoding="utf-8")
        with NamedTemporaryFile("w", encoding="utf-8", delete=False, dir=path.parent, prefix=prefix, suffix=".tmp") as tf:
            tf.write(data)
            tmp_path = Path(tf.name)
        os.replace(tmp_path, path)


def _load_index() -> dict[str, Any]:
    match_data_index_path = _match_data_index_path()
    if match_data_index_path.exists():
        with match_data_index_path.open("r", encoding="utf-8") as f:
            idx = json.load(f)
            if isinstance(idx, dict):
                return {
                    "version": int(idx.get("version", VERSION)),
                    "updatedAt": idx.get("updatedAt"),
                    "selectedDatasetId": str(idx.get("selectedDatasetId") or ""),
                    "datasets": idx.get("datasets") if isinstance(idx.get("datasets"), list) else [],
                }
    return _default_index()


def _load_dataset_doc(dataset_id: str) -> dict[str, Any] | None:
    path = _dataset_file_path(dataset_id)
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _write_dataset_doc(dataset_id: str, doc: dict[str, Any]) -> None:
    path = _dataset_file_path(dataset_id)
    bak_path = _match_datasets_dir() / f"{dataset_id}.bak.json"
    _atomic_write_json(path, bak_path, f"match_dataset_{dataset_id}_", doc)


def _write_index(doc: dict[str, Any]) -> None:
    _atomic_write_json(_match_data_index_path(), _match_data_index_bak_path(), "match_datasets_index_", doc)


def _resolve_dataset(index: dict[str, Any], requested_dataset_id: str) -> tuple[str, dict[str, Any] | None]:
    dataset_id = requested_dataset_id or str(index.get("selectedDatasetId") or "")
    if not dataset_id:
        return "", None
    return dataset_id, _load_dataset_doc(dataset_id)


def _build_columns(entity: dict[str, Any], schema: dict[str, Any]) -> list[dict[str, Any]]:
    numeric_cols = set(schema.get("numericColumns", []))
    all_cols = schema.get("allColumns", [])
    raw = entity.get("raw", {})
    metrics = entity.get("metrics", {})
    rows = []
    for col in all_cols:
        metric = metrics.get(col, {})
        rows.append(
            {
                "column": col,
                "value": raw.get(col, ""),
                "rank": metric.get("rank"),
                "percentile": metric.get("percentile"),
                "isNumeric": col in numeric_cols,
            }
        )
    return rows


@match_data_bp.route("/api/match-data/import-excel", methods=["POST"])
def import_match_data_excel():
    file = request.files.get("file")
    if file is None or not file.filename:
        return jsonify({"ok": False, "error": "missing file"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "only .xlsx is supported"}), 400

    try:
        wb = load_workbook(file, data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"invalid excel file: {exc}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return jsonify({"ok": False, "error": "excel must contain header and data rows"}), 400

    maxc = max((len(r) for r in rows), default=0)
    if maxc <= 0:
        return jsonify({"ok": False, "error": "empty sheet"}), 400

    row1 = list(rows[0]) if len(rows) > 0 else []
    row2 = list(rows[1]) if len(rows) > 1 else []
    two_header_like = False
    if len(rows) >= 3:
        has_split_header_signal = False
        for c in range(maxc):
            r1 = str(_to_cell_value(row1[c]) if c < len(row1) else "").strip()
            r2 = str(_to_cell_value(row2[c]) if c < len(row2) else "").strip()
            if (not r1 and r2) or ("/" in r1):
                has_split_header_signal = True
                break

        # Guard against misclassifying normal one-row headers that include "/"
        # in metric names (e.g. "Shots / on target"). In those files, row2 is
        # an actual data row and contains many numeric cells.
        row2_numeric_count = _count_numeric_cells(row2, maxc)
        two_header_like = has_split_header_signal and row2_numeric_count == 0

    if two_header_like:
        headers = _build_split_headers(row1, row2, maxc)
        data_rows = rows[2:]
    else:
        headers = [str(_to_cell_value(row1[c]) if c < len(row1) else "").strip() for c in range(maxc)]
        data_rows = rows[1:]

    if len(data_rows) == 0:
        return jsonify({"ok": False, "error": "excel has no data rows"}), 400

    headers_lower = [h.lower() for h in headers]
    if "team" not in headers_lower:
        return jsonify({"ok": False, "error": "missing required column: Team"}), 400
    team_idx = headers_lower.index("team")

    parsed_teams: list[dict[str, Any]] = []
    used_ids: dict[str, int] = {}
    candidate_numeric_cols = [h for h in headers if h and h.lower() != "team"]

    for row_idx, excel_row in enumerate(data_rows, start=1):
        cells = list(excel_row) + [None] * max(0, len(headers) - len(excel_row))
        raw = {}
        for i, col in enumerate(headers):
            raw[col] = _to_cell_value(cells[i] if i < len(cells) else None)

        team_name = str(raw.get(headers[team_idx], "")).strip()
        if not team_name:
            continue

        team = {
            "id": _make_id(team_name, used_ids),
            "team": team_name,
            "raw": raw,
            "metrics": {},
            "_numeric": {},
            "_rowIndex": row_idx,
        }
        parsed_teams.append(team)

    if not parsed_teams:
        return jsonify({"ok": False, "error": "no valid team rows found"}), 400

    numeric_columns, lower_better_columns = compute_player_metrics(
        parsed_teams,
        candidate_numeric_cols,
        to_float_fn=_to_float,
        is_lower_better_fn=is_lower_better_column,
    )
    for team in parsed_teams:
        team.pop("_numeric", None)
        team.pop("_rowIndex", None)

    doc = {
        "version": VERSION,
        "updatedAt": _iso_now(),
        "source": {
            "filename": file.filename,
            "sheet": ws.title,
            "rowCount": len(parsed_teams),
        },
        "schema": {
            "teamColumn": "team",
            "numericColumns": numeric_columns,
            "lowerBetterColumns": lower_better_columns,
            "allColumns": headers,
        },
        "teams": parsed_teams,
    }
    dataset_id = f"mds_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"
    doc["datasetId"] = dataset_id

    try:
        _write_dataset_doc(dataset_id, doc)
        index = _load_index()
        datasets = [d for d in index.get("datasets", []) if isinstance(d, dict)]
        datasets.insert(
            0,
            {
                "id": dataset_id,
                "name": f"{file.filename} ({doc['updatedAt'][:19]})",
                "updatedAt": doc["updatedAt"],
                "teamCount": len(parsed_teams),
                "numericColumnCount": len(numeric_columns),
                "sourceFile": file.filename,
            },
        )
        index["datasets"] = datasets
        index["selectedDatasetId"] = dataset_id
        index["updatedAt"] = doc["updatedAt"]
        _write_index(index)
    except Exception as exc:
        return jsonify({"ok": False, "error": f"write failed: {exc}"}), 500

    return jsonify(
        {
            "ok": True,
            "datasetId": dataset_id,
            "updatedAt": doc["updatedAt"],
            "teamCount": len(parsed_teams),
            "numericColumnCount": len(numeric_columns),
        }
    )


@match_data_bp.route("/api/match-data/datasets", methods=["GET"])
def get_match_datasets():
    try:
        index = _load_index()
        return jsonify(
            {
                "ok": True,
                "datasets": index.get("datasets", []),
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "updatedAt": index.get("updatedAt"),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


@match_data_bp.route("/api/match-data/datasets/<dataset_id>", methods=["DELETE"])
def delete_match_dataset(dataset_id: str):
    try:
        index = _load_index()
        datasets = [d for d in index.get("datasets", []) if isinstance(d, dict)]
        matched = next((d for d in datasets if d.get("id") == dataset_id), None)
        if matched is None:
            return jsonify({"ok": False, "error": "dataset not found"}), 404

        path = _dataset_file_path(dataset_id)
        if path.exists():
            path.unlink()

        remaining = [d for d in datasets if d.get("id") != dataset_id]
        index["datasets"] = remaining
        if index.get("selectedDatasetId") == dataset_id:
            index["selectedDatasetId"] = remaining[0].get("id") if remaining else ""
        index["updatedAt"] = _iso_now()
        _write_index(index)
        return jsonify({"ok": True, "deletedDatasetId": dataset_id, "selectedDatasetId": index.get("selectedDatasetId", ""), "datasets": remaining})
    except Exception as exc:
        return jsonify({"ok": False, "error": f"delete failed: {exc}"}), 500


@match_data_bp.route("/api/match-data/teams", methods=["GET"])
def get_match_team_list():
    try:
        index = _load_index()
        dataset_id, doc = _resolve_dataset(index, str(request.args.get("datasetId") or ""))
        if doc is None:
            return jsonify(
                {
                    "ok": True,
                    "teams": [],
                    "teamCount": 0,
                    "updatedAt": None,
                    "numericColumns": [],
                    "datasetId": dataset_id,
                    "selectedDatasetId": index.get("selectedDatasetId", ""),
                }
            )
        teams = doc.get("teams", [])
        schema = doc.get("schema", {}) if isinstance(doc.get("schema"), dict) else {}
        items = [{"id": t.get("id"), "team": t.get("team")} for t in teams if t.get("id") and t.get("team")]
        return jsonify(
            {
                "ok": True,
                "teams": items,
                "teamCount": len(items),
                "datasetId": dataset_id,
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "updatedAt": doc.get("updatedAt"),
                "numericColumns": schema.get("numericColumns", []),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500


@match_data_bp.route("/api/match-data/team/<team_id>", methods=["GET"])
def get_match_team_detail(team_id: str):
    try:
        index = _load_index()
        dataset_id, doc = _resolve_dataset(index, str(request.args.get("datasetId") or ""))
        if doc is None:
            return jsonify({"ok": False, "error": "match dataset not found"}), 404
        teams = doc.get("teams", [])
        selected = next((t for t in teams if t.get("id") == team_id), None)
        if selected is None:
            return jsonify({"ok": False, "error": "team not found"}), 404
        schema = doc.get("schema", {}) if isinstance(doc.get("schema"), dict) else {}
        return jsonify(
            {
                "ok": True,
                "team": {
                    "id": selected.get("id"),
                    "team": selected.get("team"),
                    "columns": _build_columns(selected, schema),
                },
                "datasetId": dataset_id,
                "selectedDatasetId": index.get("selectedDatasetId", ""),
                "updatedAt": doc.get("updatedAt"),
            }
        )
    except Exception as exc:
        return jsonify({"ok": False, "error": f"read failed: {exc}"}), 500
