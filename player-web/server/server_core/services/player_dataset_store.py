from __future__ import annotations

import json
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from server_core.services.ranking_service import (
    compute_player_metrics as _svc_compute_player_metrics,
    is_lower_better_column as _svc_is_lower_better_column,
    normalize_player_dataset_doc as _svc_normalize_player_dataset_doc,
)
from server_core.services.state_store import VERSION, atomic_write_json, ensure_data_dir, iso_now


APP_DIR = Path(__file__).resolve().parents[2]
DATA_DIR = APP_DIR / "data"
PLAYER_DATA_PATH = DATA_DIR / "player_dataset.json"
PLAYER_DATA_BAK_PATH = DATA_DIR / "player_dataset.json.bak"
PLAYER_DATASETS_DIR = DATA_DIR / "player_datasets"
PLAYER_DATA_INDEX_PATH = DATA_DIR / "player_datasets_index.json"
PLAYER_DATA_INDEX_BAK_PATH = DATA_DIR / "player_datasets_index.json.bak"


def ensure_player_data_dir() -> None:
    ensure_data_dir()
    PLAYER_DATASETS_DIR.mkdir(parents=True, exist_ok=True)


def to_float(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = text.replace(",", "")
        try:
            return float(text)
        except ValueError:
            return None
    return None


def to_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def normalize_header_name(value: Any) -> str:
    text = str(value or "").strip().lower().replace("_", " ")
    return " ".join(text.split())


def pick_name_column(headers: list[str]) -> tuple[int, str]:
    candidates = {"player", "name", "player name", "姓名", "球员", "球员姓名"}
    for idx, header in enumerate(headers):
        normalized = normalize_header_name(header)
        if normalized in candidates:
            return idx, header
    return -1, ""


def pick_team_column(headers: list[str]) -> tuple[int, str]:
    exact_candidates = {"team", "club", "squad", "球队", "俱乐部"}
    for idx, header in enumerate(headers):
        normalized = normalize_header_name(header)
        if normalized in exact_candidates:
            return idx, header
    keyword_candidates = ("team", "club", "squad", "球队", "俱乐部")
    for idx, header in enumerate(headers):
        normalized = normalize_header_name(header)
        if any(keyword in normalized for keyword in keyword_candidates):
            return idx, header
    return -1, ""


def is_lower_better_column(column_name: str) -> bool:
    return _svc_is_lower_better_column(column_name)


def compute_player_metrics(players: list[dict[str, Any]], candidate_numeric_cols: list[str]) -> tuple[list[str], list[str]]:
    return _svc_compute_player_metrics(
        players,
        candidate_numeric_cols,
        to_float_fn=to_float,
        is_lower_better_fn=is_lower_better_column,
    )


def normalize_player_dataset_doc(doc: dict[str, Any]) -> dict[str, Any]:
    return _svc_normalize_player_dataset_doc(
        doc,
        to_float_fn=to_float,
        is_lower_better_fn=is_lower_better_column,
    )


def make_player_id(player_name: str, used_ids: dict[str, int]) -> str:
    base = "".join(ch.lower() if ch.isalnum() else "_" for ch in player_name).strip("_")
    if not base:
        base = "player"
    if base not in used_ids:
        used_ids[base] = 1
        return base
    used_ids[base] += 1
    return f"{base}_{used_ids[base]}"


def build_player_columns(player: dict[str, Any], schema: dict[str, Any]) -> list[dict[str, Any]]:
    numeric_cols = set(schema.get("numericColumns", []))
    all_cols = schema.get("allColumns", [])
    raw = player.get("raw", {})
    metrics = player.get("metrics", {})
    rows = []
    for col in all_cols:
        metric = metrics.get(col, {})
        rows.append(
            {
                "column": col,
                "value": raw.get(col, ""),
                "rank": metric.get("rank"),
                "percentile": metric.get("percentile"),
                "isNumeric": col in numeric_cols,
            }
        )
    return rows


def default_player_index() -> dict[str, Any]:
    return {
        "version": VERSION,
        "updatedAt": None,
        "selectedDatasetId": "",
        "datasets": [],
    }


def dataset_file_path(dataset_id: str) -> Path:
    return PLAYER_DATASETS_DIR / f"{dataset_id}.json"


def load_player_doc() -> dict[str, Any] | None:
    if not PLAYER_DATA_PATH.exists():
        return None
    with PLAYER_DATA_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_dataset_doc(dataset_id: str) -> dict[str, Any] | None:
    path = dataset_file_path(dataset_id)
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def write_dataset_doc(dataset_id: str, doc: dict[str, Any]) -> None:
    ensure_player_data_dir()
    path = dataset_file_path(dataset_id)
    bak_path = PLAYER_DATASETS_DIR / f"{dataset_id}.bak.json"
    atomic_write_json(path, bak_path, f"dataset_{dataset_id}_", doc)


def write_player_doc(doc: dict[str, Any]) -> None:
    ensure_player_data_dir()
    atomic_write_json(PLAYER_DATA_PATH, PLAYER_DATA_BAK_PATH, "player_data_", doc)


def write_player_index(doc: dict[str, Any]) -> None:
    ensure_player_data_dir()
    atomic_write_json(PLAYER_DATA_INDEX_PATH, PLAYER_DATA_INDEX_BAK_PATH, "player_datasets_index_", doc)


def load_player_index() -> dict[str, Any]:
    if PLAYER_DATA_INDEX_PATH.exists():
        with PLAYER_DATA_INDEX_PATH.open("r", encoding="utf-8") as f:
            idx = json.load(f)
            if isinstance(idx, dict):
                return {
                    "version": int(idx.get("version", VERSION)),
                    "updatedAt": idx.get("updatedAt"),
                    "selectedDatasetId": str(idx.get("selectedDatasetId") or ""),
                    "datasets": idx.get("datasets") if isinstance(idx.get("datasets"), list) else [],
                }
    legacy = load_player_doc()
    if legacy and isinstance(legacy, dict) and legacy.get("players"):
        dataset_id = "legacy"
        write_dataset_doc(dataset_id, legacy)
        imported_at = legacy.get("updatedAt") or iso_now()
        idx = {
            "version": VERSION,
            "updatedAt": imported_at,
            "selectedDatasetId": dataset_id,
            "datasets": [
                {
                    "id": dataset_id,
                    "name": f"legacy-{imported_at[:19]}",
                    "updatedAt": imported_at,
                    "playerCount": len(legacy.get("players", [])),
                    "numericColumnCount": len((legacy.get("schema") or {}).get("numericColumns", [])),
                    "sourceFile": (legacy.get("source") or {}).get("filename", "legacy"),
                }
            ],
        }
        write_player_index(idx)
        return idx
    return default_player_index()


def resolve_dataset(index: dict[str, Any], requested_dataset_id: str) -> tuple[str, dict[str, Any] | None]:
    dataset_id = requested_dataset_id or str(index.get("selectedDatasetId") or "")
    if not dataset_id:
        return "", None
    return dataset_id, load_dataset_doc(dataset_id)


def import_player_excel_bytes(filename: str, payload: bytes) -> dict[str, Any]:
    if not filename:
        raise ValueError("missing file")
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("only .xlsx is supported")
    if not payload:
        raise ValueError("empty excel file")
    try:
        wb = load_workbook(BytesIO(payload), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"invalid excel file: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        raise ValueError("excel must contain header and data rows")

    headers_raw = [to_cell_value(x) for x in rows[0]]
    headers = [str(x).strip() for x in headers_raw]
    headers_lower = [h.lower() for h in headers]
    if "player" not in headers_lower:
        raise ValueError("missing required column: player")
    player_idx = headers_lower.index("player")

    parsed_players: list[dict[str, Any]] = []
    used_ids: dict[str, int] = {}
    seen_numeric_cols: set[str] = set()
    candidate_numeric_cols = []
    for header in headers:
        if not header or header.lower() == "player" or header in seen_numeric_cols:
            continue
        seen_numeric_cols.add(header)
        candidate_numeric_cols.append(header)

    for row_idx, excel_row in enumerate(rows[1:], start=1):
        cells = list(excel_row) + [None] * max(0, len(headers) - len(excel_row))
        raw = {}
        for i, col in enumerate(headers):
            value = to_cell_value(cells[i] if i < len(cells) else None)
            if col not in raw or (raw[col] == "" and value != ""):
                raw[col] = value

        player_name = str(raw.get(headers[player_idx], "")).strip()
        if not player_name:
            continue

        parsed_players.append(
            {
                "id": make_player_id(player_name, used_ids),
                "player": player_name,
                "raw": raw,
                "metrics": {},
                "_numeric": {},
                "_rowIndex": row_idx,
            }
        )

    if not parsed_players:
        raise ValueError("no valid player rows found")

    numeric_columns, lower_better_columns = compute_player_metrics(parsed_players, candidate_numeric_cols)
    for player in parsed_players:
        player.pop("_numeric", None)
        player.pop("_rowIndex", None)

    doc = {
        "version": VERSION,
        "updatedAt": iso_now(),
        "source": {
            "filename": filename,
            "sheet": ws.title,
            "rowCount": len(parsed_players),
        },
        "schema": {
            "playerColumn": "player",
            "numericColumns": numeric_columns,
            "lowerBetterColumns": lower_better_columns,
            "allColumns": headers,
        },
        "players": parsed_players,
    }
    dataset_id = f"ds_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{uuid4().hex[:6]}"
    doc["datasetId"] = dataset_id

    write_dataset_doc(dataset_id, doc)
    index = load_player_index()
    datasets = [d for d in index.get("datasets", []) if isinstance(d, dict)]
    datasets.insert(
        0,
        {
            "id": dataset_id,
            "name": f"{filename} ({doc['updatedAt'][:19]})",
            "updatedAt": doc["updatedAt"],
            "playerCount": len(parsed_players),
            "numericColumnCount": len(numeric_columns),
            "sourceFile": filename,
        },
    )
    index["datasets"] = datasets
    index["selectedDatasetId"] = dataset_id
    index["updatedAt"] = doc["updatedAt"]
    write_player_index(index)
    write_player_doc(doc)

    return {
        "datasetId": dataset_id,
        "updatedAt": doc["updatedAt"],
        "playerCount": len(parsed_players),
        "numericColumnCount": len(numeric_columns),
    }


def delete_player_dataset_doc(dataset_id: str) -> dict[str, Any]:
    index = load_player_index()
    datasets = [d for d in index.get("datasets", []) if isinstance(d, dict)]
    matched = next((d for d in datasets if d.get("id") == dataset_id), None)
    if matched is None:
        raise FileNotFoundError("dataset not found")

    path = dataset_file_path(dataset_id)
    if path.exists():
        path.unlink()

    remaining = [d for d in datasets if d.get("id") != dataset_id]
    index["datasets"] = remaining
    if index.get("selectedDatasetId") == dataset_id:
        index["selectedDatasetId"] = remaining[0].get("id") if remaining else ""
    index["updatedAt"] = iso_now()
    write_player_index(index)
    return {
        "deletedDatasetId": dataset_id,
        "selectedDatasetId": index.get("selectedDatasetId", ""),
        "datasets": remaining,
    }
